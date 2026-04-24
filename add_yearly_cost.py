"""
add_yearly_cost.py
------------------
为 Azure 原始价格表 Excel 新增 "Estimated yearly cost" 列。

逻辑：
  - 在 "Estimated upfront cost" 列之后插入新列
  - 数据服务行（Row 4 起，直到 Total 行之前）：yearly cost = monthly cost * 12
  - Total 行：yearly cost = SUM(数据服务行的年度列)
  - 其余行（Licensing Program / Billing Account / Billing Profile 等）：留空
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy
import sys
import os

INPUT_FILE  = "docs/原始价格表.xlsx"
OUTPUT_FILE = "docs/原始价格表_with_yearly.xlsx"


def find_header_row(ws):
    """找到含 'Estimated monthly cost' 的标题行号"""
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and "Estimated monthly cost" in row:
            return i
    return None


def find_total_row(ws, header_row):
    """找到 Total 所在行号"""
    for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        if row and "Total" in row:
            return i
    return None


def col_letter(n):
    """将列号（1-indexed）转为字母，如 1→A, 7→G"""
    result = ""
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def copy_cell_style(src_cell, dst_cell):
    """复制单元格样式"""
    if src_cell.has_style:
        dst_cell.font      = copy(src_cell.font)
        dst_cell.fill      = copy(src_cell.fill)
        dst_cell.border    = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def process_sheet(ws):
    header_row = find_header_row(ws)
    if header_row is None:
        print(f"  ⚠ 未在 Sheet '{ws.title}' 中找到标题行，跳过。")
        return

    total_row = find_total_row(ws, header_row)
    if total_row is None:
        print(f"  ⚠ 未在 Sheet '{ws.title}' 中找到 Total 行，跳过。")
        return

    # 找 monthly cost 列和 upfront cost 列
    header_values = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    try:
        monthly_col  = header_values.index("Estimated monthly cost") + 1   # 1-indexed
        upfront_col  = header_values.index("Estimated upfront cost") + 1
    except ValueError:
        print(f"  ⚠ 未找到必要的列名，跳过 Sheet '{ws.title}'。")
        return

    # 新列紧接在 upfront_col 之后
    yearly_col = upfront_col + 1

    print(f"  Sheet: '{ws.title}'")
    print(f"  标题行: {header_row}，Total 行: {total_row}")
    print(f"  Monthly cost 列: {col_letter(monthly_col)} ({monthly_col})")
    print(f"  Upfront cost 列: {col_letter(upfront_col)} ({upfront_col})")
    print(f"  新增 Yearly cost 列: {col_letter(yearly_col)} ({yearly_col})")

    # --- 插入新列 ---
    ws.insert_cols(yearly_col)

    # 更新列字母（insert 之后 upfront_col 之后的列号不变，yearly_col 已是新列）
    monthly_letter = col_letter(monthly_col)
    yearly_letter  = col_letter(yearly_col)

    # 1. 写标题
    header_cell = ws.cell(header_row, yearly_col, "Estimated yearly cost")
    # 复制 upfront cost 标题的样式
    src_header = ws.cell(header_row, upfront_col)
    copy_cell_style(src_header, header_cell)
    header_cell.font = Font(
        name=src_header.font.name or "Calibri",
        bold=True,
        size=src_header.font.size or 11,
        color=src_header.font.color.rgb if src_header.font.color and src_header.font.color.type == "rgb" else "000000",
    )

    # 数据行范围：header_row+1 到 total_row-1（不含 Total 行本身）
    data_start = header_row + 1
    data_end   = total_row - 1

    # 2. 数据服务行：=monthly * 12（用公式）
    for r in range(data_start, data_end + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, 6)]
        # 跳过 Licensing/Billing 等元信息行（D 列含关键词但 A 列为空）
        # 判断：若该行 monthly cost 列有数值或 Total 行 => 写公式
        monthly_val = ws.cell(r, monthly_col).value
        if monthly_val is not None and isinstance(monthly_val, (int, float)):
            cell = ws.cell(r, yearly_col)
            cell.value = f"={monthly_letter}{r}*12"
            # 复制同行 monthly 列的样式
            copy_cell_style(ws.cell(r, monthly_col), cell)
            cell.number_format = '#,##0.00'
        else:
            # 留空（Licensing Program / Billing Account 等行）
            ws.cell(r, yearly_col).value = None

    # 3. Total 行：=SUM(yearly_data_range)
    total_cell = ws.cell(total_row, yearly_col)
    sum_range  = f"{yearly_letter}{data_start}:{yearly_letter}{data_end}"
    total_cell.value = f"=SUM({sum_range})"
    # 复制 Total 行 monthly 列的样式
    copy_cell_style(ws.cell(total_row, monthly_col), total_cell)
    total_cell.number_format = '#,##0.00'
    total_cell.font = Font(bold=True, name="Calibri", size=11)

    # 4. 调整新列宽度
    ws.column_dimensions[yearly_letter].width = 22

    print(f"  ✅ 已写入 yearly cost 列，数据行 {data_start}-{data_end}，Total: {total_row}")


def main():
    if not os.path.exists(INPUT_FILE):
        print(f"❌ 文件不存在: {INPUT_FILE}")
        sys.exit(1)

    print(f"📂 读取: {INPUT_FILE}")
    wb = openpyxl.load_workbook(INPUT_FILE)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        process_sheet(ws)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 已保存到: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
