"""Azure Migrate assessment orchestration."""

import datetime
import io
import re
import time
from typing import Any, Callable, Dict, List, Optional

import requests

from azure.arm import azure_arm_list, azure_arm_request, register_azure_provider
from budget.parser import _format_usd, parse_annual_budget_usd
from budget.tier import (
    budget_target_range,
    ensure_csv_disk_columns,
    format_budget_target_range,
    get_machine_ids_for_tier,
    jitter_csv_performance,
    learn_tier_machine_selections,
    load_builtin_csv_template,
    load_tier_cache,
    prefix_csv_server_names,
    save_tier_cache,
    _safe_csv_prefix,
)
from config import (
    AZURE_MIGRATE_API_VERSION,
    AZURE_MIGRATE_V2_API_VERSION,
    AZURE_MIGRATE_DEFAULT_TARGET_LOCATION,
    AZURE_MIGRATE_PROJECTS_API_VERSION,
    AZURE_MIGRATE_REPORT_API_VERSION,
    AZURE_OFFAZURE_API_VERSION,
    AZURE_RESOURCE_API_VERSION,
)

def _date_prefix():
    """返回当前日期前缀，如 0225"""
    return datetime.date.today().strftime("%m%d")


# ──────────────────────────────────────────────
# 全自动 POE：MSAL 登录与 Azure ARM API
# ──────────────────────────────────────────────

def _safe_azure_name(value: str, fallback: str, suffix: str = "", max_len: int = 60) -> str:
    base = re.sub(r"[^A-Za-z0-9-]+", "-", value or "").strip("-")
    base = re.sub(r"-+", "-", base)
    if not base:
        base = fallback
    reserve = len(suffix)
    return f"{base[: max_len - reserve].strip('-')}{suffix}".strip("-")


AZURE_MIGRATE_PROJECT_LOCATIONS = {
    "centralus", "westeurope", "uksouth", "ukwest", "northeurope", "westus2",
    "southeastasia", "eastasia", "centralindia", "southindia", "canadacentral",
    "australiasoutheast", "japanwest", "japaneast", "brazilsouth", "koreacentral",
    "koreasouth", "francecentral", "switzerlandnorth", "australiaeast", "uaenorth",
    "southafricanorth", "germanywestcentral", "norwayeast", "jioindiawest",
    "swedencentral", "qatarcentral", "polandcentral", "italynorth", "israelcentral",
    "spaincentral", "mexicocentral", "newzealandnorth", "indonesiacentral",
    "malaysiawest", "chilecentral", "austriaeast", "belgiumcentral", "denmarkeast",
}

AZURE_MIGRATE_LOCATION_ALIASES = {
    "westus": "westus2",
    "west us": "westus2",
    "east us 2": "eastus2",
    "eastus 2": "eastus2",
}


def _normalize_azure_location(location: str) -> str:
    return re.sub(r"\s+", "", (location or "").strip().lower())


def resolve_migrate_project_location(subscription_id: str, resource_group: str, token: str) -> str:
    resource_group_payload = azure_arm_request(
        "GET",
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}?api-version={AZURE_RESOURCE_API_VERSION}",
        token,
    )
    normalized = _normalize_azure_location(resource_group_payload.get("location", ""))
    normalized = AZURE_MIGRATE_LOCATION_ALIASES.get(normalized, normalized)
    if normalized in AZURE_MIGRATE_PROJECT_LOCATIONS:
        return normalized
    return "westus2"



def _extract_sas_url(payload: Any) -> Optional[str]:
    if isinstance(payload, str):
        if payload.startswith("http") and "sig=" in payload:
            return payload
        return None
    if isinstance(payload, dict):
        for value in payload.values():
            found = _extract_sas_url(value)
            if found:
                return found
    if isinstance(payload, list):
        for value in payload:
            found = _extract_sas_url(value)
            if found:
                return found
    return None


def _arm_path_with_api_version(path_or_id: str, api_version: str) -> str:
    separator = "&" if "?" in path_or_id else "?"
    return f"{path_or_id}{separator}api-version={api_version}"


def _resource_id(subscription_id: str, resource_group: str, provider_path: str) -> str:
    return f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}/{provider_path}"


def _migrate_project_path(subscription_id: str, resource_group: str, project_name: str) -> str:
    return (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Migrate/migrateProjects/{project_name}"
        f"?api-version={AZURE_MIGRATE_PROJECTS_API_VERSION}"
    )


def _migrate_solution_path(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    solution_name: str,
) -> str:
    return (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Migrate/migrateProjects/{project_name}"
        f"/solutions/{solution_name}?api-version={AZURE_MIGRATE_PROJECTS_API_VERSION}"
    )


def _migrate_solution_id(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    solution_name: str,
) -> str:
    return _resource_id(
        subscription_id,
        resource_group,
        f"providers/Microsoft.Migrate/migrateProjects/{project_name}/solutions/{solution_name}",
    )


def _servers_solution_details(extended_details: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    details = {
        "dependencyEnabledMachines": "0",
        "machinesHavingSqlServers": "0",
        "machinesHavingWebServers": "0",
        "serversOnLinux": "0",
        "serversOnWindows": "0",
        "serversOnOther": "0",
    }
    if extended_details:
        details.update(extended_details)
    return {
        "assessmentCount": 0,
        "groupCount": 0,
        "extendedDetails": details,
    }


def register_migrate_tool(subscription_id: str, resource_group: str, project_name: str, tool: str, token: str) -> None:
    azure_arm_request(
        "POST",
        (
            f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
            f"/providers/Microsoft.Migrate/migrateProjects/{project_name}"
            f"/registerTool?api-version={AZURE_MIGRATE_PROJECTS_API_VERSION}"
        ),
        token,
        {"tool": tool},
    )


def put_migrate_solution(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    solution_name: str,
    properties: Dict[str, Any],
    token: str,
) -> Dict[str, Any]:
    return azure_arm_request(
        "PUT",
        _migrate_solution_path(subscription_id, resource_group, project_name, solution_name),
        token,
        {"properties": properties},
    )


def ensure_migrate_solution(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    solution_name: str,
    properties: Dict[str, Any],
    token: str,
    progress: Optional[Callable[[str], None]] = None,
) -> Dict[str, Any]:
    path = _migrate_solution_path(subscription_id, resource_group, project_name, solution_name)
    existing = _try_get_existing_resource(path, token)
    if existing and existing.get("properties", {}).get("details"):
        if progress:
            progress(f"  ✅ Solution 已存在且 details 完整，复用: {solution_name}")
        return existing
    result = put_migrate_solution(subscription_id, resource_group, project_name, solution_name, properties, token)
    if progress:
        progress(f"  ✅ Solution 已补齐: {solution_name}")
    return result


def ensure_portal_menu_solutions(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    master_site_id: str,
    token: str,
    progress: Callable[[str], None],
) -> None:
    """补齐 Azure Portal 项目菜单 blade 会枚举的默认 solution，避免前端读取 undefined.details。"""
    default_solutions = [
        (
            "Servers-Discovery-ServerDiscovery",
            {
                "tool": "ServerDiscovery",
                "purpose": "Discovery",
                "goal": "Servers",
                "status": "Inactive",
                "details": _servers_solution_details({"masterSiteId": master_site_id}),
            },
        ),
        (
            "Servers-Migration-ServerMigration",
            {
                "tool": "ServerMigration",
                "purpose": "Migration",
                "goal": "Servers",
                "status": "Active",
                "details": _servers_solution_details(),
            },
        ),
        (
            "Servers-Migration-ServerMigration_DataReplication",
            {
                "tool": "ServerMigration_DataReplication",
                "purpose": "Migration",
                "goal": "Servers",
                "status": "Inactive",
                "details": _servers_solution_details(),
            },
        ),
    ]
    for solution_name, properties in default_solutions:
        ensure_migrate_solution(
            subscription_id,
            resource_group,
            project_name,
            solution_name,
            properties,
            token,
            progress,
        )


def refresh_migrate_project_summary(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    token: str,
) -> None:
    azure_arm_request(
        "POST",
        (
            f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
            f"/providers/Microsoft.Migrate/migrateProjects/{project_name}"
            f"/refreshSummary?api-version={AZURE_MIGRATE_PROJECTS_API_VERSION}"
        ),
        token,
        {"goal": "Servers"},
        poll_lro=False,
    )


def _server_summary_count(migrate_project: Dict[str, Any]) -> int:
    servers = migrate_project.get("properties", {}).get("summary", {}).get("servers", {})
    direct_count = int(servers.get("discoveredCount") or 0)
    extended = servers.get("extendedSummary") or {}
    microsoft_count = int(extended.get("microsoftMachinesCount") or 0)
    return max(direct_count, microsoft_count)


def wait_for_portal_inventory_summary(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    expected_machine_count: int,
    token: str,
    progress: Callable[[str], None],
    timeout_seconds: int = 300,
) -> int:
    """等待 migrateProject summary 反映 Import CSV 库存，这是 Portal 全部库存 blade 使用的项目汇总层。"""
    project_path = _migrate_project_path(subscription_id, resource_group, project_name)
    deadline = time.time() + timeout_seconds
    attempt = 0
    last_count = 0

    while time.time() < deadline:
        attempt += 1
        try:
            refresh_migrate_project_summary(subscription_id, resource_group, project_name, token)
        except Exception:
            pass

        project = azure_arm_request("GET", project_path, token, poll_lro=False)
        last_count = _server_summary_count(project)
        if last_count >= expected_machine_count:
            return last_count

        progress(
            f"等待 Azure Portal 全部库存汇总刷新，当前 {last_count}/{expected_machine_count} 台"
            f"（第 {attempt} 次检查）"
        )
        time.sleep(20)

    raise TimeoutError(
        "Azure Portal 全部库存汇总未刷新到预期数量。"
        f"当前 {last_count}/{expected_machine_count} 台；"
        "请检查 ServerDiscovery_Import solution 与 Import Site 关联。"
    )


def _format_import_job_error(job: Dict[str, Any]) -> str:
    props = job.get("properties", {}) if isinstance(job, dict) else {}
    summary = props.get("errorSummary") if isinstance(props, dict) else {}
    parts = []
    if isinstance(summary, dict):
        error_count = summary.get("errorCount")
        warning_count = summary.get("warningCount")
        if error_count is not None:
            parts.append(f"errors={error_count}")
        if warning_count is not None:
            parts.append(f"warnings={warning_count}")
        errors = summary.get("errors")
        if isinstance(errors, list) and errors:
            preview = "; ".join(str(item) for item in errors[:5])
            parts.append(f"details={preview}")
    result = props.get("jobResult") or job.get("status")
    if result:
        parts.insert(0, f"jobResult={result}")
    return "；".join(parts) if parts else str(job)[:800]


def _dedupe_machines_by_discovery_arm_id(machines: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    deduped: List[Dict[str, Any]] = []
    seen = set()
    for machine in machines:
        props = machine.get("properties", {})
        key = str(props.get("discoveryMachineArmId") or machine.get("id") or machine.get("name") or "").lower()
        if not key or key in seen:
            continue
        seen.add(key)
        deduped.append(machine)
    return deduped


def wait_for_import_site_import(
    subscription_id: str,
    resource_group: str,
    site_name: str,
    token: str,
    progress: Callable[[str], None],
    job_arm_id: Optional[str] = None,
    timeout_seconds: int = 900,
) -> List[Dict[str, Any]]:
    """等待 OffAzure import site 完成 CSV 解析，并返回导入到 import site 的机器。"""
    machines_path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.OffAzure/importSites/{site_name}/machines"
        f"?api-version={AZURE_OFFAZURE_API_VERSION}"
    )
    jobs_path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.OffAzure/importSites/{site_name}/importJobs"
        f"?api-version={AZURE_OFFAZURE_API_VERSION}"
    )
    job_path = (
        _arm_path_with_api_version(job_arm_id, AZURE_OFFAZURE_API_VERSION)
        if job_arm_id and (job_arm_id.startswith("/") or job_arm_id.startswith("http"))
        else None
    )
    deadline = time.time() + timeout_seconds
    attempt = 0
    last_job: Dict[str, Any] = {}

    while time.time() < deadline:
        attempt += 1
        try:
            machines = azure_arm_list(machines_path, token)
            if machines:
                return machines
        except Exception:
            pass

        job: Dict[str, Any] = {}
        if job_path:
            try:
                job = azure_arm_request("GET", job_path, token, poll_lro=False)
            except Exception:
                job = {}
        if not job:
            try:
                jobs = azure_arm_list(jobs_path, token)
                if jobs:
                    job = jobs[-1]
            except Exception:
                job = {}
        if job:
            last_job = job
            props = job.get("properties", {})
            result = str(props.get("jobResult") or job.get("status") or "Unknown").strip()
            imported_count = props.get("numberOfMachinesImported")
            if result in {"Completed", "CompletedWithWarnings"}:
                machines = azure_arm_list(machines_path, token)
                if machines:
                    return machines
            if result in {"Failed", "CompletedWithErrors"}:
                raise RuntimeError(f"Azure Migrate CSV 导入失败：{_format_import_job_error(job)}")
            suffix = f"，已导入 {imported_count} 台" if imported_count is not None else ""
            progress(f"服务器清单导入任务状态：{result}{suffix}（第 {attempt} 次检查）")
        else:
            progress(f"等待 Azure Migrate 创建服务器清单导入任务...（第 {attempt} 次检查）")
        time.sleep(15)

    detail = f"最后一次任务状态：{_format_import_job_error(last_job)}" if last_job else "未查询到导入任务。"
    raise TimeoutError(f"等待 Azure Migrate 导入服务器清单超时（15 分钟）。{detail}")


def wait_for_project_machines(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    collector_name: str,
    token: str,
    progress: Callable[[str], None],
    site_name: Optional[str] = None,
    timeout_seconds: int = 600,
) -> List[Dict[str, Any]]:
    """等待 Azure Migrate 导入完成，通过直接查询 machines 列表来判断。"""
    machines_path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Migrate/assessmentProjects/{project_name}"
        f"/machines?api-version={AZURE_MIGRATE_API_VERSION}"
    )
    project_path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Migrate/assessmentProjects/{project_name}"
        f"?api-version={AZURE_MIGRATE_API_VERSION}"
    )
    deadline = time.time() + timeout_seconds
    attempt = 0

    def _filter_current_import(machines: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        if not site_name:
            return _dedupe_machines_by_discovery_arm_id(machines)
        marker = f"/importsites/{site_name}".lower()
        return _dedupe_machines_by_discovery_arm_id([
            machine for machine in machines
            if marker in str(machine.get("properties", {}).get("discoveryMachineArmId", "")).lower()
        ])

    while time.time() < deadline:
        attempt += 1
        # 先尝试直接列出 machines
        try:
            all_machines = azure_arm_list(machines_path, token)
            machines = _filter_current_import(all_machines)
            if machines:
                return machines
        except Exception:
            pass
        # 也检查 project 的 numberOfMachines
        try:
            project = azure_arm_request("GET", project_path, token)
            machine_count = project.get("properties", {}).get("numberOfMachines", 0) or 0
            if machine_count > 0:
                machines = _filter_current_import(azure_arm_list(machines_path, token))
                if machines:
                    return machines
        except Exception:
            pass
        progress(f"服务器清单仍在导入中，继续等待 Azure Migrate 发现结果...（第 {attempt} 次检查）")
        time.sleep(15)
    raise TimeoutError("等待 Azure Migrate 导入服务器清单超时（10 分钟），请在 Azure Portal 检查导入状态。")


def list_imported_machines(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    collector_name: str,
    token: str,
) -> List[Dict[str, Any]]:
    candidate_paths = [
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}/providers/Microsoft.Migrate/assessmentProjects/{project_name}/machines?api-version={AZURE_MIGRATE_API_VERSION}",
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}/providers/Microsoft.Migrate/assessmentProjects/{project_name}/importcollectors/{collector_name}/machines?api-version={AZURE_MIGRATE_API_VERSION}",
    ]
    last_error = None
    for path in candidate_paths:
        try:
            machines = azure_arm_list(path, token)
            if machines:
                return machines
        except Exception as exc:
            last_error = exc
    if last_error:
        raise RuntimeError(f"无法读取 Azure Migrate 已导入服务器列表：{last_error}")
    return []


def wait_for_assessment_complete(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    group_name: str,
    assessment_name: str,
    token: str,
    progress: Callable[[str], None],
    timeout_seconds: int = 600,
) -> Dict[str, Any]:
    path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Migrate/assessmentProjects/{project_name}"
        f"/groups/{group_name}/assessments/{assessment_name}"
        f"?api-version={AZURE_MIGRATE_API_VERSION}"
    )
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        assessment = azure_arm_request("GET", path, token)
        props = assessment.get("properties", {})
        status = props.get("status", "Unknown")
        stage = props.get("stage", "Unknown")
        provisioning_state = props.get("provisioningState", "Unknown")
        if status == "Completed":
            return assessment
        if status in {"Invalid", "OutOfSync", "OutDated", "Deleted"}:
            raise RuntimeError(f"Azure Migrate 评估状态异常：{status}")
        if provisioning_state in {"Failed", "Canceled"}:
            raise RuntimeError(f"Azure Migrate 评估资源状态异常：{provisioning_state}")
        progress(f"评估仍在计算中，当前状态：{status}，阶段：{stage}，资源状态：{provisioning_state}")
        time.sleep(15)
    raise TimeoutError("等待 Azure Migrate 评估完成超时，请稍后在 Azure Portal 检查评估结果。")



def _assessment_cost_component(assessment: Dict[str, Any], component_name: str) -> float:
    props = assessment.get("properties", {})
    for component in props.get("costComponents") or []:
        if str(component.get("name") or "").lower() == component_name.lower():
            try:
                return float(component.get("value") or 0)
            except (TypeError, ValueError):
                return 0.0
    return 0.0


def assessment_monthly_total_cost(assessment: Dict[str, Any]) -> float:
    props = assessment.get("properties", {})

    def _num(name: str) -> float:
        try:
            return float(props.get(name) or 0)
        except (TypeError, ValueError):
            return 0.0

    return (
        _num("monthlyComputeCost")
        + _num("monthlyStorageCost")
        + _num("monthlyBandwidthCost")
        + _assessment_cost_component(assessment, "MonthlySecurityCost")
    )


def _format_usd(value: Optional[float]) -> str:
    if value is None:
        return "未填写"
    return f"${value:,.2f}"


def _assessment_settings_snapshot(settings: Dict[str, Any]) -> Dict[str, Any]:
    keys = [
        "azureLocation",
        "sizingCriterion",
        "reservedInstance",
        "azureHybridUseBenefit",
        "linuxAzureHybridUseBenefit",
        "azureSecurityOfferingType",
        "scalingFactor",
        "discountPercentage",
    ]
    return {key: settings.get(key) for key in keys}


# Azure Migrate 评估地区名称映射（解决方案文档中的区域名 → API azureLocation 值）
_REGION_NAME_TO_LOCATION = {
    "east us": "EastUs", "east us 2": "EastUs2", "west us": "WestUs",
    "west us 2": "WestUs2", "west us 3": "WestUs3", "central us": "CentralUs",
    "north central us": "NorthCentralUs", "south central us": "SouthCentralUs",
    "west europe": "WestEurope", "north europe": "NorthEurope",
    "southeast asia": "SoutheastAsia", "east asia": "EastAsia",
    "japan east": "JapanEast", "japan west": "JapanWest",
    "australia east": "AustraliaEast", "australia southeast": "AustraliaSoutheast",
    "uk south": "UKSouth", "uk west": "UKWest",
    "canada central": "CanadaCentral", "canada east": "CanadaEast",
    "korea central": "KoreaCentral", "korea south": "KoreaSouth",
    "france central": "FranceCentral", "germany west central": "GermanyWestCentral",
    "switzerland north": "SwitzerlandNorth", "norway east": "NorwayEast",
    "brazil south": "BrazilSouth", "south africa north": "SouthAfricaNorth",
    "uae north": "UAENorth", "india central": "CentralIndia",
    "india south": "SouthIndia", "india west": "WestIndia",
    "sweden central": "SwedenCentral", "qatar central": "QatarCentral",
}


def _extract_dominant_region(solution_text: str) -> Optional[str]:
    """
    从解决方案文本中提取出现频率最高的 Azure 区域，返回对应的 azureLocation API 值。
    如果无法识别任何区域，返回 None。
    """
    if not solution_text:
        return None
    text_lower = solution_text.lower()
    region_counts: Dict[str, int] = {}
    # 按名称长度降序匹配，避免 "east us" 匹配到 "east us 2" 的情况
    sorted_regions = sorted(_REGION_NAME_TO_LOCATION.keys(), key=len, reverse=True)
    for region_name in sorted_regions:
        count = text_lower.count(region_name)
        if count > 0:
            location_val = _REGION_NAME_TO_LOCATION[region_name]
            region_counts[location_val] = region_counts.get(location_val, 0) + count
    if not region_counts:
        return None
    # 返回出现次数最多的区域
    return max(region_counts, key=region_counts.get)


def _build_assessment_body(target_location: Optional[str] = None) -> Dict[str, Any]:
    return {
        "properties": {
            "groupType": "Import",
            "assessmentType": "MachineAssessment",
            "azureLocation": target_location or AZURE_MIGRATE_DEFAULT_TARGET_LOCATION,
            "azureOfferCode": "MSAZR0003P",
            "azurePricingTier": "Standard",
            "azureStorageRedundancy": "LocallyRedundant",
            "scalingFactor": 1.3,
            "percentile": "Percentile95",
            "timeRange": "Day",
            "currency": "USD",
            "azureHybridUseBenefit": "Yes",
            "linuxAzureHybridUseBenefit": "Yes",
            "azureSecurityOfferingType": "MDC",
            "discountPercentage": 0,
            "sizingCriterion": "PerformanceBased",
            "azureDiskTypes": ["PremiumSSD", "StandardSSD", "StandardHDD"],
            "azureVmFamilies": [
                "Dv2_series", "Dv3_series", "DSv2_series", "Dsv3_series", "Ev3_series",
                "Esv3_series", "F_series", "Fs_series", "Fsv2_series", "M_series", "D_series",
                "DS_series", "H_series", "Lsv2_series",
            ],
            "vmUptime": {"daysPerMonth": 31, "hoursPerDay": 24},
            "reservedInstance": "RI3Year",
            "stage": "InProgress",
        },
    }


def _build_v2_arg_query(import_site_id: str, machine_ids: List[str]) -> str:
    """构建新版评估使用的 Azure Resource Graph 查询，用于筛选指定服务器。"""
    machine_id_list = ", ".join(f'"{mid.lower()}"' for mid in machine_ids)
    return (
        "migrateresources\n"
        '        | where type in~ ("microsoft.offazure/vmwaresites/machines", '
        '"microsoft.offazure/hypervsites/machines", '
        '"microsoft.offazure/serversites/machines", '
        '"microsoft.offazure/importsites/machines")\n'
        f'        | where id has "{import_site_id}"\n'
        "        | extend type=tolower(type)\n"
        "        | extend id = tolower(id)\n"
        f"        | where id in~ ({machine_id_list})"
    )


def _build_assessment_body_v2(
    import_site_id: str,
    machine_ids: List[str],
    target_location: Optional[str] = None,
    customer_name: str = "",
) -> Dict[str, Any]:
    """构建新版 Azure VM 评估的请求体（project-level，无需 group）。"""
    # 客户级确定性差异化：初始评估参数不同，避免同 tier 客户最终金额完全一致。
    # discount 是直接价格乘数；scalingFactor 会影响 SKU 推荐边界。
    scaling_factor = 1.3
    discount_percentage = 0.0
    if customer_name:
        import hashlib as _hashlib
        import random as _rng
        seed = int(_hashlib.md5(customer_name.encode()).hexdigest(), 16) % (2**32)
        rng = _rng.Random(seed)
        scaling_factor = round(rng.uniform(1.15, 1.45), 2)
        discount_percentage = round(rng.uniform(0, 6), 2)

    return {
        "properties": {
            "details": {
                "assessmentType": "MachineAssessment",
            },
            "scope": {
                "scopeType": "AzureResourceGraphQuery",
                "azureResourceGraphQuery": _build_v2_arg_query(import_site_id, machine_ids),
            },
            "settings": {
                "azureLocation": target_location or AZURE_MIGRATE_DEFAULT_TARGET_LOCATION,
                "azurePricingTier": "Standard",
                "azureStorageRedundancy": "LocallyRedundant",
                "scalingFactor": scaling_factor,
                "currency": "USD",
                "azureHybridUseBenefit": "Yes",
                "linuxAzureHybridUseBenefit": "Yes",
                "azureSecurityOfferingType": "MDC",
                "discountPercentage": discount_percentage,
                "sizingCriterion": "PerformanceBased",
                "azureDiskTypes": ["Premium", "StandardSSD", "PremiumV2"],
                "azureVmFamilies": [
                    "Dadsv5_series", "Dasv4_series", "Dasv5_series",
                    "Ddsv4_series", "Ddsv5_series", "Ddv5_series",
                    "Dsv4_series", "Dsv5_series", "Dv5_series",
                    "Edsv5_series", "Edv5_series", "Esv5_series", "Ev5_series",
                    "Eadsv5_series", "Easv4_series", "Easv5_series",
                    "Ebdsv5_series", "Ebsv5_series", "Edsv4_series", "Esv4_series",
                    "Fs_series", "Fsv2_series",
                    "Lasv3_series", "Lsv2_series", "Lsv3_series",
                    "M_series", "Mdsv2_series", "Msv2_series", "Mv2_series",
                    "Av2_series", "Dav4_series", "Ddv4_series", "Dv4_series",
                    "Eav4_series", "Edv4_series", "Ev4_series", "F_series",
                ],
                "azureVmSecurityOptions": ["Standard", "TVM"],
                "billingSettings": {"licensingProgram": "Retail", "subscriptionId": None},
                "environmentType": "Production",
                "performanceData": {
                    "percentile": "Percentile95",
                    "timeRange": "Day",
                },
                "savingsSettings": {
                    "azureOfferCode": "MSAZR0003P",
                    "savingsOptions": "RI3Year",
                },
                "vmUptime": {"daysPerMonth": 31, "hoursPerDay": 24},
            },
        },
    }


def wait_for_assessment_complete_v2(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    assessment_name: str,
    token: str,
    progress: Callable[[str], None],
    timeout_seconds: int = 600,
) -> Dict[str, Any]:
    """等待新版 project-level 评估完成（status 在 properties.details.status）。"""
    path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Migrate/assessmentprojects/{project_name}"
        f"/assessments/{assessment_name}"
        f"?api-version={AZURE_MIGRATE_V2_API_VERSION}"
    )
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        assessment = azure_arm_request("GET", path, token)
        details = assessment.get("properties", {}).get("details", {})
        status = details.get("status", "Unknown")
        if status == "Completed":
            return assessment
        if status in {"Invalid", "OutOfSync", "OutDated", "Deleted"}:
            raise RuntimeError(f"Azure Migrate 评估状态异常：{status}")
        progress(f"评估仍在计算中，当前状态：{status}")
        time.sleep(15)
    raise TimeoutError("等待 Azure Migrate 评估完成超时，请稍后在 Azure Portal 检查评估结果。")


def _extract_machine_monthly_cost(machine: Dict[str, Any]) -> float:
    """从单台 V2 assessed machine 中提取月成本。

    V2 API 成本路径：recommendations[].totalCost[].costDetail[]
    只取 name=TotalMonthlyCost 的条目，其余为组件明细（会重复计算）。
    """
    mp = machine.get("properties", {})

    for rec in mp.get("recommendations", []):
        rec_total_cost = rec.get("totalCost")
        if isinstance(rec_total_cost, list):
            for cost_group in rec_total_cost:
                if isinstance(cost_group, dict):
                    for detail in cost_group.get("costDetail", []):
                        if str(detail.get("name", "")).lower() == "totalmonthlycost":
                            try:
                                return float(detail.get("value") or 0)
                            except (TypeError, ValueError):
                                pass

        # 回退：sku 级别
        for sku in rec.get("skus", []):
            details = sku.get("details", {})
            sku_total_cost = details.get("totalCost")
            if isinstance(sku_total_cost, list):
                for cost_group in sku_total_cost:
                    if isinstance(cost_group, dict):
                        for detail in cost_group.get("costDetail", []):
                            if str(detail.get("name", "")).lower() == "totalmonthlycost":
                                try:
                                    return float(detail.get("value") or 0)
                                except (TypeError, ValueError):
                                    pass

    return 0.0


def assessment_monthly_total_cost_v2(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    assessment_name: str,
    token: str,
) -> float:
    """从新版评估的 assessedMachines 聚合月总成本。

    V2 API 成本路径：recommendations[].totalCost[].costDetail[].value
    """
    machines_path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Migrate/assessmentprojects/{project_name}"
        f"/assessments/{assessment_name}/assessedMachines"
        f"?api-version={AZURE_MIGRATE_V2_API_VERSION}"
    )
    machines = azure_arm_list(machines_path, token)
    total = 0.0
    for machine in machines:
        total += _extract_machine_monthly_cost(machine)
    return total


def download_assessment_report_v2(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    assessment_name: str,
    token: str,
) -> bytes:
    """下载新版 project-level 评估报告。"""
    download_url_path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Migrate/assessmentprojects/{project_name}"
        f"/assessments/{assessment_name}/downloadUrl"
        f"?api-version={AZURE_MIGRATE_V2_API_VERSION}"
    )
    payload = azure_arm_request("POST", download_url_path, token, {}, poll_lro=False)
    report_url = payload.get("assessmentReportUrl")
    if not report_url:
        raise RuntimeError(f"Azure Migrate 未返回评估报告下载地址：{payload}")
    response = requests.get(report_url, timeout=180)
    if response.status_code >= 400:
        raise RuntimeError(f"下载 Azure Migrate 评估报告失败：{response.status_code} {response.text[:800]}")
    if not response.content:
        raise RuntimeError("Azure Migrate 评估报告为空。")
    return response.content


def _put_assessment_v2_and_wait(
    assessment_path: str,
    assessment_body: Dict[str, Any],
    token: str,
    subscription_id: str,
    resource_group: str,
    project_name: str,
    assessment_name: str,
    progress: Callable[[str], None],
    timeout_seconds: int = 600,
) -> Dict[str, Any]:
    """PUT 新版评估更新并等待计算完成，自动处理 405 冲突。"""
    deadline = time.time() + timeout_seconds
    attempt = 0
    while time.time() < deadline:
        attempt += 1
        try:
            azure_arm_request("PUT", assessment_path, token, assessment_body)
            break
        except Exception as e:
            msg = str(e)
            if "405" in msg or "under computation" in msg.lower() or "Cannot be updated" in msg:
                if attempt == 1:
                    progress(f"  ⏳ 评估仍在计算中，等待完成后再更新...")
                time.sleep(15)
                continue
            raise

    return wait_for_assessment_complete_v2(
        subscription_id, resource_group, project_name, assessment_name, token, progress
    )


def tune_assessment_to_budget_v2(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    assessment_name: str,
    assessment_path: str,
    assessment_body: Dict[str, Any],
    annual_budget: Optional[float],
    token: str,
    progress: Callable[[str], None],
    customer_name: str = "",
) -> tuple[float, List[Dict[str, Any]], bool]:
    """针对新版评估的预算调优，返回 (monthly_cost, history, met)。"""
    history: List[Dict[str, Any]] = []
    range_info = budget_target_range(annual_budget)
    target_min = range_info["target_min"]
    target_max = range_info["target_max"]
    target_mid = range_info["target_mid"]
    target_max_exclusive = bool(range_info["target_max_exclusive"])
    target_anchor = target_mid
    if target_min is not None and target_max is not None and customer_name:
        import hashlib as _hashlib
        import random as _rng
        seed = int(_hashlib.md5(customer_name.encode()).hexdigest(), 16) % (2**32)
        rng = _rng.Random(seed)
        span = max(float(target_max) - float(target_min), 0)
        # 留出上下边界缓冲，让自动调优落在区间内部且不同客户锚点不同。
        target_anchor = float(target_min) + span * rng.uniform(0.35, 0.75)

    def _above_target(annual_total: float) -> bool:
        if target_max is None:
            return False
        return annual_total >= target_max if target_max_exclusive else annual_total > target_max

    def _in_target_range(annual_total: float) -> bool:
        if target_min is None or target_max is None:
            return True
        below_upper = annual_total < target_max if target_max_exclusive else annual_total <= target_max
        return target_min <= annual_total and below_upper

    def _get_monthly_cost() -> float:
        return assessment_monthly_total_cost_v2(
            subscription_id, resource_group, project_name, assessment_name, token
        )

    def _record(round_name: str, action: str, monthly_total: float, met: bool) -> None:
        annual_total = monthly_total * 12
        history.append({
            "round": round_name,
            "action": action,
            "monthly_total": monthly_total,
            "annual_total": annual_total,
            "target_annual": annual_budget,
            "target_min": target_min,
            "target_max": target_max,
            "target_anchor": target_anchor,
            "met_target": met,
        })

    if annual_budget is None or annual_budget <= 0:
        monthly_total = _get_monthly_cost()
        progress(
            "未填写可解析的预估年消耗，跳过价格校准。"
            f"当前 Azure Migrate 年化估算：{_format_usd(monthly_total * 12)}"
        )
        _record("initial", "未填写可解析预算，未调整评估设置", monthly_total, True)
        return monthly_total, history, True

    monthly_total = _get_monthly_cost()
    annual_total = monthly_total * 12
    progress(
        "Azure Migrate 当前年化估算："
        f"{_format_usd(annual_total)}；目标区间：{format_budget_target_range(range_info)}"
    )
    if target_anchor is not None:
        progress(f"客户专属调优目标锚点：{_format_usd(float(target_anchor))}")
    if _in_target_range(annual_total):
        _record("initial", "初始评估已在目标区间内", monthly_total, True)
        return monthly_total, history, True
    direction = "高于" if _above_target(annual_total) else "低于"
    _record("initial", f"初始评估{direction}目标区间", monthly_total, False)

    settings = assessment_body["properties"]["settings"]

    def _next_patch(annual_total: float) -> tuple[str, Dict[str, Any]]:
        if target_max is not None and target_anchor is not None and _above_target(annual_total):
            # --- 成本高于目标：需要降低 ---
            current_discount = float(settings.get("discountPercentage") or 0)
            current_discount_factor = max(1 - current_discount / 100, 0.01)
            undiscounted_annual = annual_total / current_discount_factor
            required_discount = 100 * (1 - (target_anchor / max(undiscounted_annual, 1)))
            next_discount = min(max(current_discount, required_discount), 99.0)
            if next_discount > current_discount + 0.1:
                return (
                    f"设置折扣为 {next_discount:.2f}%，控制年化估算落入客户专属目标区间",
                    {"discountPercentage": round(next_discount, 2)},
                )
            current_factor = float(settings.get("scalingFactor") or 1.3)
            if current_factor > 1.0:
                next_factor = max(1.0, current_factor * 0.8)
                return (
                    f"降低舒适因子到 {next_factor:.2f}，控制年化估算落入目标区间",
                    {"scalingFactor": round(next_factor, 2)},
                )
            if settings.get("azureSecurityOfferingType") != "NO":
                return ("关闭安全成本估算，控制年化估算落入目标区间", {"azureSecurityOfferingType": "NO"})
            return ("已达到自动降价边界", {})

        if target_anchor is None:
            return ("未设置预算，不调整", {})

        # --- 成本低于目标：需要提高 ---
        # 计算差距比例，决定是否组合多个杠杆一次性应用
        gap_ratio = float(target_anchor) / max(annual_total, 1)  # >1 means below target
        combined_patch: Dict[str, Any] = {}
        actions: list[str] = []

        # 杠杆1：取消折扣
        current_discount = float(settings.get("discountPercentage") or 0)
        if current_discount > 0:
            combined_patch["discountPercentage"] = 0
            actions.append("取消折扣")

        # 杠杆2：关闭 Hybrid Benefit（OS 许可成本计入估算）
        if settings.get("azureHybridUseBenefit") != "No" or settings.get("linuxAzureHybridUseBenefit") != "No":
            combined_patch["azureHybridUseBenefit"] = "No"
            combined_patch["linuxAzureHybridUseBenefit"] = "No"
            actions.append("关闭 Hybrid Benefit")

        # 杠杆3：切换预留实例为按量计费（影响最大，约 2-3 倍）
        if settings.get("savingsSettings", {}).get("savingsOptions") != "None":
            combined_patch["savingsSettings"] = {"azureOfferCode": "MSAZR0003P", "savingsOptions": "None"}
            actions.append("切换为按量计费")

        # 如果差距大（>2x），一次性应用所有可用杠杆 + 拉满舒适因子
        if gap_ratio > 2.0 and len(actions) > 1:
            current_factor = float(settings.get("scalingFactor") or 1.3)
            if current_factor < 2.0:
                combined_patch["scalingFactor"] = 2.0
                actions.append("舒适因子拉满到 2.0")
            return ("、".join(actions) + "，大幅提高年化估算", combined_patch)

        # 差距较小时，逐个应用杠杆（优先大杠杆）
        if settings.get("savingsSettings", {}).get("savingsOptions") != "None":
            return ("切换为按量计费以提高年化估算", {"savingsSettings": {"azureOfferCode": "MSAZR0003P", "savingsOptions": "None"}})
        if settings.get("azureHybridUseBenefit") != "No" or settings.get("linuxAzureHybridUseBenefit") != "No":
            return (
                "关闭 Azure Hybrid Benefit，把 OS 许可成本计入估算",
                {"azureHybridUseBenefit": "No", "linuxAzureHybridUseBenefit": "No"},
            )
        if current_discount > 0:
            return ("取消折扣以提高年化估算", {"discountPercentage": 0})

        # 杠杆4：提高舒适因子（上限 2.0）—— 差距大时直接拉满
        current_factor = float(settings.get("scalingFactor") or 1.3)
        if current_factor < 2.0:
            next_factor = 2.0 if gap_ratio > 1.5 else min(current_factor * min(gap_ratio, 1.5), 2.0)
            if next_factor > current_factor + 0.01:
                return (f"提高舒适因子到 {next_factor:.2f}", {"scalingFactor": round(next_factor, 2)})

        # 杠杆5：切换性能百分位到 99%（选择更大 VM 规格）
        perf = settings.get("performanceData", {})
        if perf.get("percentile") != "Percentile99":
            return ("切换到 99 百分位性能数据（选择更大 VM）", {"performanceData": {**perf, "percentile": "Percentile99"}})

        # 杠杆6：切换为按本地规格评估（使用实际硬件配置，通常选择更大 VM）
        if settings.get("sizingCriterion") != "AsOnPremises":
            return ("切换为按本地配置评估（不按性能缩放，通常得到更大 VM）", {"sizingCriterion": "AsOnPremises"})

        # 杠杆7：仅保留 Premium 存储（费用更高）
        disk_types = settings.get("azureDiskTypes", [])
        if "StandardSSD" in disk_types and len(disk_types) > 1:
            premium_only = [t for t in disk_types if t != "StandardSSD"]
            return ("移除标准 SSD 存储选项，仅保留高级存储", {"azureDiskTypes": premium_only})

        return ("已达到自动提价边界", {})

    for round_index in range(1, 6):
        round_name = f"round-{round_index}"
        action, patch = _next_patch(annual_total)
        if not patch:
            progress(f"评估年化估算仍不在目标区间内，{action}。")
            break
        progress(f"评估年化估算不在目标区间，开始自动调整（{round_name}）：{action}")
        settings.update(patch)
        _put_assessment_v2_and_wait(
            assessment_path, assessment_body, token,
            subscription_id, resource_group, project_name, assessment_name,
            progress,
        )
        monthly_total = _get_monthly_cost()
        annual_total = monthly_total * 12
        met = _in_target_range(annual_total)
        progress(
            f"{round_name} 重新计算完成：月估算 {_format_usd(monthly_total)}，"
            f"年化 {_format_usd(annual_total)}，目标区间 {format_budget_target_range(range_info)}"
        )
        _record(round_name, action, monthly_total, met)
        if met:
            return monthly_total, history, True

    progress(
        f"已自动调整 {round_index} 轮，但 Azure Migrate 年化估算仍未落入用户预估年消耗的档次区间"
        f"（{format_budget_target_range(range_info)}）；"
        "请到 Azure Portal 的评估设置中手动调整后重新导出。"
    )
    return monthly_total, history, False


def _put_assessment_and_wait(
    assessment_path: str,
    assessment_body: Dict[str, Any],
    token: str,
    subscription_id: str,
    resource_group: str,
    project_name: str,
    group_name: str,
    assessment_name: str,
    progress: Callable[[str], None],
    timeout_seconds: int = 600,
) -> Dict[str, Any]:
    """PUT 评估更新并等待计算完成，自动处理 405 冲突。"""
    deadline = time.time() + timeout_seconds
    attempt = 0
    while time.time() < deadline:
        attempt += 1
        try:
            azure_arm_request("PUT", assessment_path, token, assessment_body)
            break
        except Exception as e:
            msg = str(e)
            if "405" in msg or "under computation" in msg.lower() or "Cannot be updated" in msg:
                if attempt == 1:
                    progress(f"  ⏳ 评估仍在计算中，等待完成后再更新...")
                time.sleep(15)
                continue
            raise

    return wait_for_assessment_complete(
        subscription_id, resource_group, project_name, group_name, assessment_name, token, progress
    )


def tune_assessment_to_budget(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    group_name: str,
    assessment_name: str,
    assessment_path: str,
    assessment_body: Dict[str, Any],
    assessment: Dict[str, Any],
    annual_budget: Optional[float],
    token: str,
    progress: Callable[[str], None],
) -> tuple[Dict[str, Any], List[Dict[str, Any]], bool]:
    history: List[Dict[str, Any]] = []
    range_info = budget_target_range(annual_budget)
    target_min = range_info["target_min"]
    target_max = range_info["target_max"]
    target_mid = range_info["target_mid"]
    target_max_exclusive = bool(range_info["target_max_exclusive"])

    def _above_target(annual_total: float) -> bool:
        if target_max is None:
            return False
        return annual_total >= target_max if target_max_exclusive else annual_total > target_max

    def _record(round_name: str, action: str, current: Dict[str, Any], met: bool) -> None:
        monthly_total = assessment_monthly_total_cost(current)
        annual_total = monthly_total * 12
        history.append({
            "round": round_name,
            "action": action,
            "monthly_total": monthly_total,
            "annual_total": annual_total,
            "target_annual": annual_budget,
            "target_min": target_min,
            "target_max": target_max,
            "met_target": met,
            "settings": _assessment_settings_snapshot(assessment_body["properties"]),
        })

    def _in_target_range(current: Dict[str, Any]) -> bool:
        if target_min is None or target_max is None:
            return True
        annual_total = assessment_monthly_total_cost(current) * 12
        below_upper = annual_total < target_max if target_max_exclusive else annual_total <= target_max
        return target_min <= annual_total and below_upper

    def _next_patch(annual_total: float) -> tuple[str, Dict[str, Any]]:
        settings = assessment_body["properties"]
        if target_max is not None and target_mid is not None and _above_target(annual_total):
            current_discount = float(settings.get("discountPercentage") or 0)
            current_discount_factor = max(1 - current_discount / 100, 0.01)
            undiscounted_annual = annual_total / current_discount_factor
            required_discount = 100 * (1 - (target_mid / max(undiscounted_annual, 1)))
            next_discount = min(max(current_discount, required_discount), 99.0)
            if next_discount > current_discount + 0.1:
                return (
                    f"设置折扣为 {next_discount:.2f}%，控制年化估算落入目标区间",
                    {"discountPercentage": round(next_discount, 2)},
                )
            current_factor = float(settings.get("scalingFactor") or 1.3)
            if current_factor > 1.0:
                next_factor = max(1.0, current_factor * 0.8)
                return (
                    f"降低舒适因子到 {next_factor:.2f}，控制年化估算落入目标区间",
                    {"scalingFactor": round(next_factor, 2)},
                )
            if settings.get("azureSecurityOfferingType") != "NO":
                return ("关闭安全成本估算，控制年化估算落入目标区间", {"azureSecurityOfferingType": "NO"})
            return ("已达到自动降价边界", {})

        if target_mid is None:
            return ("未设置预算，不调整", {})

        current_discount = float(settings.get("discountPercentage") or 0)
        if current_discount > 0:
            return ("取消折扣以提高年化估算", {"discountPercentage": 0})
        if settings.get("azureHybridUseBenefit") != "No" or settings.get("linuxAzureHybridUseBenefit") != "No":
            return (
                "关闭 Azure Hybrid Benefit，把 OS 许可成本计入估算",
                {"azureHybridUseBenefit": "No", "linuxAzureHybridUseBenefit": "No"},
            )
        current_factor = float(settings.get("scalingFactor") or 1.3)
        factor_ratio = min(max(target_mid / max(annual_total, 1), 1.05), 1.35)
        next_factor = min(current_factor * factor_ratio, 5.0)
        if next_factor > current_factor + 0.01:
            return (f"提高舒适因子到 {next_factor:.2f}", {"scalingFactor": round(next_factor, 2)})
        if settings.get("reservedInstance") != "None":
            return ("切换为按量计费以提高年化估算", {"reservedInstance": "None"})
        return ("已达到自动提价边界", {})

    if annual_budget is None or annual_budget <= 0:
        monthly_total = assessment_monthly_total_cost(assessment)
        progress(
            "未填写可解析的预估年消耗，跳过价格校准。"
            f"当前 Azure Migrate 年化估算：{_format_usd(monthly_total * 12)}"
        )
        _record("initial", "未填写可解析预算，未调整评估设置", assessment, True)
        return assessment, history, True

    monthly_total = assessment_monthly_total_cost(assessment)
    annual_total = monthly_total * 12
    progress(
        "Azure Migrate 当前年化估算："
        f"{_format_usd(annual_total)}；目标区间：{format_budget_target_range(range_info)}"
    )
    if _in_target_range(assessment):
        _record("initial", "初始 Portal 默认评估已在目标区间内", assessment, True)
        return assessment, history, True
    direction = "高于" if _above_target(annual_total) else "低于"
    _record("initial", f"初始 Portal 默认评估{direction}目标区间", assessment, False)

    current_props = assessment_body["properties"]
    # 如有必要，调整存储冗余级别以微调成本
    for redundancy in ["LocallyRedundant", "GeoRedundant"]:
        if redundancy == current_props.get("azureStorageRedundancy", ""):
            continue
        test_body = {**assessment_body, "properties": {**assessment_body["properties"], "azureStorageRedundancy": redundancy}}
        try:
            test_result = _put_assessment_and_wait(
                assessment_path, test_body, token,
                subscription_id, resource_group, project_name, group_name, assessment_name,
                progress,
            )
        except Exception as e:
            progress(f"  ⚠️ 存储冗余调整失败：{e}")
            continue
        test_monthly = assessment_monthly_total_cost(test_result)
        test_annual = test_monthly * 12
        if target_min and test_annual < target_min:
            continue
        if _above_target(test_annual):
            continue
        assessment_body = test_body
        assessment = test_result
        annual_total = test_annual
        progress(f"  调整为存储冗余 {redundancy}: 年化 {_format_usd(test_annual)} ✓")
        _record("storage-redundancy", f"调整为存储冗余 {redundancy}", assessment, True)
        return assessment, history, True

    for round_index in range(1, 4):
        round_name = f"round-{round_index}"
        action, patch = _next_patch(annual_total)
        if not patch:
            progress(f"评估年化估算仍不在目标区间内，{action}。")
            break
        progress(f"评估年化估算不在目标区间，开始自动调整（{round_name}）：{action}")
        assessment_body["properties"].update(patch)
        assessment_body["properties"]["stage"] = "InProgress"
        assessment = _put_assessment_and_wait(
            assessment_path, assessment_body, token,
            subscription_id, resource_group, project_name, group_name, assessment_name,
            progress,
        )
        monthly_total = assessment_monthly_total_cost(assessment)
        annual_total = monthly_total * 12
        met = _in_target_range(assessment)
        progress(
            f"{round_name} 重新计算完成：月估算 {_format_usd(monthly_total)}，"
            f"年化 {_format_usd(annual_total)}，目标区间 {format_budget_target_range(range_info)}"
        )
        _record(round_name, action, assessment, met)
        if met:
            return assessment, history, True

    progress(
        f"已自动调整 3 轮，但 Azure Migrate 年化估算仍未落入用户预估年消耗的档次区间"
        f"（{format_budget_target_range(range_info)}）；"
        "请到 Azure Portal 的评估设置中手动调整后重新导出。"
    )
    return assessment, history, False


def download_assessment_report(
    subscription_id: str,
    resource_group: str,
    project_name: str,
    group_name: str,
    assessment_name: str,
    token: str,
) -> bytes:
    download_url_path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Migrate/assessmentProjects/{project_name}"
        f"/groups/{group_name}/assessments/{assessment_name}/downloadUrl"
        f"?api-version={AZURE_MIGRATE_REPORT_API_VERSION}"
    )
    payload = azure_arm_request("POST", download_url_path, token, poll_lro=False)
    report_url = payload.get("assessmentReportUrl")
    if not report_url:
        raise RuntimeError(f"Azure Migrate 未返回评估报告下载地址：{payload}")

    response = requests.get(report_url, timeout=180)
    if response.status_code >= 400:
        raise RuntimeError(f"下载 Azure Migrate 评估报告失败：{response.status_code} {response.text[:800]}")
    if not response.content:
        raise RuntimeError("Azure Migrate 评估报告为空。")
    return response.content


def _try_get_existing_resource(path: str, token: str) -> Optional[Dict[str, Any]]:
    """尝试 GET 某资源，如果存在返回 dict，不存在返回 None。"""
    try:
        result = azure_arm_request("GET", path, token)
        if result and result.get("id"):
            return result
    except Exception:
        pass
    return None


def wait_for_group_machine_membership(
    group_path: str,
    token: str,
    expected_machine_count: int,
    progress: Callable[[str], None],
    timeout_seconds: int = 600,
) -> Dict[str, Any]:
    """等待 updateMachines 完成，并返回最新评估组信息。"""
    deadline = time.time() + timeout_seconds
    attempt = 0
    last_group: Dict[str, Any] = {}

    while time.time() < deadline:
        attempt += 1
        last_group = azure_arm_request("GET", group_path, token, poll_lro=False)
        props = last_group.get("properties", {})
        machine_count = int(props.get("machineCount") or 0)
        provisioning_state = props.get("provisioningState", "Unknown")
        supported_types = sorted(
            str(item).strip()
            for item in (props.get("supportedAssessmentTypes") or [])
            if str(item).strip()
        )
        supported_label = ", ".join(supported_types) if supported_types else "未返回"

        if machine_count >= expected_machine_count and provisioning_state not in {"Failed", "Canceled"}:
            progress(f"  ℹ️ 评估组已包含 {machine_count} 台服务器；支持类型: {supported_label}")
            return last_group
        if provisioning_state in {"Failed", "Canceled"}:
            raise RuntimeError(f"评估组更新失败，资源状态：{provisioning_state}")

        progress(
            f"评估组仍在关联服务器，当前 {machine_count}/{expected_machine_count} 台；"
            f"支持类型: {supported_label}（第 {attempt} 次检查）"
        )
        time.sleep(10)

    props = last_group.get("properties", {})
    raise TimeoutError(
        "等待评估组关联服务器超时。"
        f"当前 machineCount={props.get('machineCount')}，"
        f"supportedAssessmentTypes={props.get('supportedAssessmentTypes')}"
    )


def run_azure_migrate_assessment(
    token: str,
    subscription_id: str,
    resource_group: str,
    account_name: str,
    assessment_name: str,
    annual_budget_text: Optional[str],
    progress: Callable[[str], None],
    target_location: Optional[str] = None,
) -> Dict[str, Any]:
    # ── 加载内置 CSV 模板，给服务器名加上客户前缀 ──
    progress("加载内置服务器清单模板...")
    csv_text_raw = load_builtin_csv_template()
    safe_prefix = _safe_csv_prefix(account_name)
    csv_text = prefix_csv_server_names(csv_text_raw, safe_prefix)
    csv_text = ensure_csv_disk_columns(csv_text)
    csv_text = jitter_csv_performance(csv_text, account_name)
    csv_bytes = csv_text.encode("utf-8-sig")
    progress(f"  ✅ 已为所有服务器名添加前缀: {safe_prefix}-")
    progress("  ✅ 已补齐服务器磁盘字段，确保 Azure Migrate 生成磁盘评估")
    progress("  ✅ 已应用客户专属性能参数微调")

    annual_budget = parse_annual_budget_usd(annual_budget_text)
    budget_range = budget_target_range(annual_budget)
    tier = int(budget_range["tier"])
    progress(
        f"客户预估年消耗: {_format_usd(annual_budget)}，"
        f"匹配规模档位: {_format_usd(float(tier))}，"
        f"目标区间: {format_budget_target_range(budget_range)}"
    )

    safe_base = _safe_azure_name(account_name, f"poe-{_date_prefix()}", max_len=36).lower()
    run_suffix = str(int(time.time()))
    short_run_suffix = run_suffix[-6:]
    project_name = _safe_azure_name(safe_base, "poe", "project", 55)
    site_name = _safe_azure_name(safe_base, "poe", f"site{short_run_suffix}", 24)
    master_site_name = _safe_azure_name(safe_base, "poe", "masterSite", 55)
    collector_name = _safe_azure_name(safe_base, "poe", f"collector-{short_run_suffix}", 55)
    group_name = _safe_azure_name(safe_base, "poe", f"group-{run_suffix}", 55)
    assessment_resource_name = _safe_azure_name(assessment_name, "poe-assessment", max_len=55)
    project_location = resolve_migrate_project_location(subscription_id, resource_group, token)
    migrate_project_id = _resource_id(
        subscription_id,
        resource_group,
        f"providers/Microsoft.Migrate/migrateProjects/{project_name}",
    )
    assessment_project_id = _resource_id(
        subscription_id,
        resource_group,
        f"providers/Microsoft.Migrate/assessmentProjects/{project_name}",
    )
    master_site_id = _resource_id(
        subscription_id,
        resource_group,
        f"providers/Microsoft.OffAzure/masterSites/{master_site_name}",
    )
    import_site_id = _resource_id(
        subscription_id,
        resource_group,
        f"providers/Microsoft.OffAzure/importSites/{site_name}",
    )

    progress("注册 Microsoft.Migrate 与 Microsoft.OffAzure 资源提供程序...")
    register_azure_provider(subscription_id, "Microsoft.Migrate", token)
    register_azure_provider(subscription_id, "Microsoft.OffAzure", token)

    # ── Step 1: 创建 migrateProject（Portal 可见） ──
    progress(f"创建 Azure Migrate 项目（区域：{project_location}）...")
    migrate_project_path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Migrate/migrateProjects/{project_name}"
        f"?api-version={AZURE_MIGRATE_PROJECTS_API_VERSION}"
    )
    existing_mp = _try_get_existing_resource(migrate_project_path, token)
    if existing_mp:
        mp_id = existing_mp.get("id", project_name)
        progress(f"成功复用 Azure Migrate 项目：{project_name}")
    else:
        migrate_project_body = {
            "properties": {},
            "location": project_location,
            "tags": {"Migrate Project": project_name, "createdBy": "POE Workflow"},
            "identity": {"type": "SystemAssigned"},
        }
        try:
            mp_result = azure_arm_request("PUT", migrate_project_path, token, migrate_project_body)
        except Exception:
            migrate_project_body.pop("identity", None)
            mp_result = azure_arm_request("PUT", migrate_project_path, token, migrate_project_body)
        mp_id = mp_result.get("id", project_name)
        progress(f"成功创建 Azure Migrate 项目：{project_name}")

    # ── Step 2: 注册 Portal 同款 Discovery Import 与 Assessment 工具 ──
    progress("注册 ServerDiscovery_Import 与 ServerAssessment 工具...")
    for tool in ("ServerDiscovery_Import", "ServerAssessment"):
        try:
            register_migrate_tool(subscription_id, resource_group, project_name, tool, token)
            progress(f"  ✅ 工具已注册: {tool}")
        except Exception:
            progress(f"  ℹ️ 工具可能已注册: {tool}")

    # ── Step 3: 创建 ServerAssessment Solution ──
    assessment_solution_name = "Servers-Assessment-ServerAssessment"
    assessment_solution_path = _migrate_solution_path(
        subscription_id, resource_group, project_name, assessment_solution_name
    )
    existing_sol = _try_get_existing_resource(assessment_solution_path, token)
    if existing_sol:
        assessment_solution_id = existing_sol.get("id", "")
        progress(f"成功复用评估 Solution：{assessment_solution_name}")
    else:
        sol_result = put_migrate_solution(
            subscription_id,
            resource_group,
            project_name,
            assessment_solution_name,
            {
                "tool": "ServerAssessment",
                "purpose": "Assessment",
                "goal": "Servers",
                "status": "Active",
                "details": _servers_solution_details({
                    "projectId": assessment_project_id,
                    "avsAssessment": "0",
                    "azureSqlAssessment": "0",
                    "azureVmAssessment": "0",
                    "azureWebAppAssessment": "0",
                    "businessCaseCount": "0",
                }),
            },
            token,
        )
        assessment_solution_id = sol_result.get("id", "")
        progress(f"成功创建评估 Solution：{assessment_solution_name}")

    # ── Step 4: 创建 assessmentProject 并关联 Solution ──
    progress("创建 Assessment Project...")
    ap_path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Migrate/assessmentProjects/{project_name}"
        f"?api-version={AZURE_MIGRATE_API_VERSION}"
    )
    existing_ap = _try_get_existing_resource(ap_path, token)
    ap_body = {
        "kind": "Migrate",
        "properties": {
            "projectStatus": "Active",
            "assessmentSolutionId": assessment_solution_id,
            "publicNetworkAccess": "Enabled",
        },
        "location": project_location,
        "tags": {"createdBy": "POE Workflow"},
    }
    if existing_ap and str(existing_ap.get("kind") or "").lower() == "migrate":
        ap_id = existing_ap.get("id", project_name)
        progress(f"成功复用 Assessment Project：{project_name}")
    else:
        ap_result = azure_arm_request("PUT", ap_path, token, ap_body)
        ap_id = ap_result.get("id", project_name)
        progress(f"成功创建 Assessment Project：{project_name}")

    # Portal 的评估 blade 通过 Assessment Solution 的 projectId 找到 assessmentProject。
    assessment_solution = put_migrate_solution(
        subscription_id,
        resource_group,
        project_name,
        assessment_solution_name,
        {
            "tool": "ServerAssessment",
            "purpose": "Assessment",
            "goal": "Servers",
            "status": "Active",
            "details": _servers_solution_details({
                "projectId": assessment_project_id,
                "avsAssessment": "0",
                "azureSqlAssessment": "0",
                "azureVmAssessment": "0",
                "azureWebAppAssessment": "0",
                "businessCaseCount": "0",
            }),
        },
        token,
    )
    assessment_solution_id = assessment_solution.get("id", assessment_solution_id)
    progress("  ✅ Assessment Solution 已关联 assessmentProject")

    # ── Step 5: 创建 Portal Discovery Import 链路（Master Site + Discovery Solution + Import Site） ──
    progress("创建 Portal 可识别的 Discovery Import 链路...")
    discovery_solution_name = "Servers-Discovery-ServerDiscovery_Import"
    discovery_solution_id = _migrate_solution_id(
        subscription_id, resource_group, project_name, discovery_solution_name
    )
    master_site_path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.OffAzure/masterSites/{master_site_name}"
        f"?api-version={AZURE_OFFAZURE_API_VERSION}"
    )
    existing_master_site = _try_get_existing_resource(master_site_path, token)
    existing_sites = []
    if existing_master_site:
        existing_sites = existing_master_site.get("properties", {}).get("sites") or []
        # 等待 masterSite provisioningState 变为 terminal 状态后再修改
        prov_state = existing_master_site.get("properties", {}).get("provisioningState", "")
        if prov_state and prov_state not in ("Succeeded", "Failed", "Canceled", ""):
            progress(f"  ⏳ Master Site 正在操作中（{prov_state}），等待完成...")
            for _ in range(40):  # 最多等 ~200 秒
                time.sleep(5)
                check = _try_get_existing_resource(master_site_path, token)
                if not check:
                    break
                prov_state = check.get("properties", {}).get("provisioningState", "Succeeded")
                if prov_state in ("Succeeded", "Failed", "Canceled", ""):
                    break
            progress(f"  ✅ Master Site 已就绪（{prov_state}）")
    master_site_result = azure_arm_request("PUT", master_site_path, token, {
        "kind": "Migrate",
        "location": project_location,
        "tags": {"Migrate Project": project_name, "createdBy": "POE Workflow"},
        "properties": {
            "allowMultipleSites": True,
            "publicNetworkAccess": "Enabled",
            "sites": existing_sites,
        },
    })
    master_site_id = master_site_result.get("id", master_site_id)
    progress(f"成功创建 Master Site：{master_site_name}")
    ensure_portal_menu_solutions(
        subscription_id,
        resource_group,
        project_name,
        master_site_id,
        token,
        progress,
    )

    put_migrate_solution(
        subscription_id,
        resource_group,
        project_name,
        discovery_solution_name,
        {
            "tool": "ServerDiscovery_Import",
            "purpose": "Discovery",
            "goal": "Servers",
            "status": "Inactive",
            "details": _servers_solution_details({
                "importSiteId": import_site_id,
            }),
        },
        token,
    )
    progress("  ✅ Discovery Import Solution 已关联 importSite")

    # ── Step 6: 创建 Import Site ──
    progress("创建 Import Site...")
    site_path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.OffAzure/importSites/{site_name}"
        f"?api-version={AZURE_OFFAZURE_API_VERSION}"
    )
    existing_site = _try_get_existing_resource(site_path, token)
    if existing_site:
        site_id = existing_site.get("id", site_name)
        progress(f"成功复用 Import Site：{site_name}")
    else:
        site_result = azure_arm_request("PUT", site_path, token, {
            "location": project_location,
            "properties": {
                "masterSiteId": master_site_id,
                "discoverySolutionId": discovery_solution_id,
            },
        })
        site_id = site_result.get("id", site_name)
        progress(f"成功创建 Import Site：{site_name}")

    normalized_sites = {str(site).lower(): site for site in existing_sites}
    normalized_sites.setdefault(import_site_id.lower(), import_site_id)
    azure_arm_request("PUT", master_site_path, token, {
        "kind": "Migrate",
        "location": project_location,
        "tags": {"Migrate Project": project_name, "createdBy": "POE Workflow"},
        "properties": {
            "allowMultipleSites": True,
            "publicNetworkAccess": "Enabled",
            "sites": list(normalized_sites.values()),
        },
    })
    progress("  ✅ Master Site 已关联 Import Site")

    # ── Step 7: 创建 Import Collector，关联 Import Site 到 Assessment Project ──
    progress("创建 Import Collector 关联 Import Site 到 Assessment Project...")
    collector_path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Migrate/assessmentProjects/{project_name}"
        f"/importcollectors/{collector_name}?api-version={AZURE_MIGRATE_API_VERSION}"
    )
    discovery_site_id = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.OffAzure/importSites/{site_name}"
    )
    coll_result = azure_arm_request("PUT", collector_path, token, {
        "properties": {"discoverySiteId": discovery_site_id}
    })
    progress(f"成功创建 Import Collector：{collector_name}")

    # ── Step 8: 检查当前 Import Site 是否已有库存 —— 有则跳过 CSV 导入 ──
    site_machines_path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.OffAzure/importSites/{site_name}"
        f"/machines?api-version={AZURE_OFFAZURE_API_VERSION}"
    )
    try:
        existing_site_machines = azure_arm_list(site_machines_path, token)
    except Exception:
        existing_site_machines = []

    if existing_site_machines:
        progress(f"  ℹ️ Import Site 已有 {len(existing_site_machines)} 台服务器库存，跳过 CSV 重新导入")
        imported_site_machines = existing_site_machines
        portal_inventory_count = len(existing_site_machines)
    else:
        # ── 获取 SAS URL 并上传 CSV ──
        progress("获取 CSV 上传地址并上传服务器清单...")
        import_uri_path = (
            f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
            f"/providers/Microsoft.OffAzure/importSites/{site_name}/importUri"
            f"?api-version={AZURE_OFFAZURE_API_VERSION}"
        )
        import_uri_payload = azure_arm_request("POST", import_uri_path, token, {})
        sas_url = _extract_sas_url(import_uri_payload)
        if not sas_url:
            raise RuntimeError(f"Azure 未返回可用的 CSV 上传 SAS URL：{import_uri_payload}")
        import_job_arm_id = import_uri_payload.get("jobArmId") if isinstance(import_uri_payload, dict) else None

        upload_response = requests.put(
            sas_url,
            data=csv_bytes,
            headers={"x-ms-blob-type": "BlockBlob", "Content-Type": "text/csv"},
            timeout=180,
        )
        if upload_response.status_code >= 400:
            raise RuntimeError(f"上传 CSV 到 Azure Migrate 失败：{upload_response.status_code} {upload_response.text[:800]}")
        progress(f"  ✅ CSV 已上传（{len(csv_bytes)} 字节）")

        # ── Step 9: 触发 Import Job（回传 importUri 返回的 SasUriResponse） ──
        progress("触发 Import Job 导入服务器清单...")
        import_trigger_body = dict(import_uri_payload) if isinstance(import_uri_payload, dict) else {}
        import_trigger_body["uri"] = sas_url
        if import_job_arm_id:
            import_trigger_body["jobArmId"] = import_job_arm_id
        job_result = azure_arm_request("POST", import_uri_path, token, import_trigger_body)
        import_job_arm_id = (
            job_result.get("jobArmId")
            or import_job_arm_id
            or job_result.get("id")
            if isinstance(job_result, dict)
            else import_job_arm_id
        )
        progress("成功触发 Import Job")

        # ── Step 10: 等待 OffAzure Import Site 完成 CSV 解析 ──
        imported_site_machines = wait_for_import_site_import(
            subscription_id,
            resource_group,
            site_name,
            token,
            progress,
            job_arm_id=import_job_arm_id,
        )
        progress(f"  ✅ Import Site 已导入 {len(imported_site_machines)} 台服务器")

        # Import Collector 在导入完成后再 PUT 一次，触发 assessmentProject 拉取刚导入的机器。
        azure_arm_request("PUT", collector_path, token, {
            "properties": {"discoverySiteId": discovery_site_id}
        })
        progress("  ✅ Import Collector 已刷新同步")

        portal_inventory_count = wait_for_portal_inventory_summary(
            subscription_id,
            resource_group,
            project_name,
            len(imported_site_machines),
            token,
            progress,
        )
        progress(f"  ✅ Azure Portal 全部库存汇总已刷新: {portal_inventory_count} 台服务器")

    # ── Step 11: 等待 Assessment Project 可读取机器 ──
    machines = wait_for_project_machines(
        subscription_id,
        resource_group,
        project_name,
        collector_name,
        token,
        progress,
        site_name=site_name,
    )
    all_machine_ids = [machine.get("id") for machine in machines if machine.get("id")]
    if not all_machine_ids:
        raise RuntimeError("Azure Migrate 未返回可加入评估的服务器，请检查 CSV 导入结果。")
    progress(f"  ✅ 已发现 {len(all_machine_ids)} 台服务器")

    # ── Step 12: 分规模学习 — 确定当前规模应选哪些服务器 ──
    cache = load_tier_cache()
    tier_cached = bool(cache.get("tiers", {}).get(str(tier)))

    if not tier_cached:
        progress(f"规模 {_format_usd(float(tier))} 尚未学习，开始全量评估学习...")
        learning_group_name = _safe_azure_name(safe_base, "poe", f"learn-{run_suffix}", 55)
        learning_group_path = (
            f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
            f"/providers/Microsoft.Migrate/assessmentProjects/{project_name}"
            f"/groups/{learning_group_name}?api-version={AZURE_MIGRATE_API_VERSION}"
        )
        azure_arm_request("PUT", learning_group_path, token, {
            "properties": {"groupType": "Import"},
            "eTag": "",
        })
        progress(f"  ✅ 已创建学习评估组: {learning_group_name}")

        learn_update_path = (
            f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
            f"/providers/Microsoft.Migrate/assessmentProjects/{project_name}"
            f"/groups/{learning_group_name}/updateMachines?api-version={AZURE_MIGRATE_API_VERSION}"
        )
        azure_arm_request("POST", learn_update_path, token, {
            "eTag": "*",
            "properties": {"operationType": "Add", "machines": all_machine_ids},
        })
        wait_for_group_machine_membership(
            learning_group_path, token,
            expected_machine_count=len(all_machine_ids),
            progress=progress,
        )
        progress(f"  ✅ 学习评估组已关联 {len(all_machine_ids)} 台服务器")

        learning_assess_name = _safe_azure_name("learning", "poe-assess", max_len=55)
        learning_assess_path = (
            f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
            f"/providers/Microsoft.Migrate/assessmentProjects/{project_name}"
            f"/groups/{learning_group_name}/assessments/{learning_assess_name}"
            f"?api-version={AZURE_MIGRATE_API_VERSION}"
        )
        learning_body = _build_assessment_body()
        azure_arm_request("PUT", learning_assess_path, token, learning_body)
        learning_assessment = wait_for_assessment_complete(
            subscription_id, resource_group, project_name,
            learning_group_name, learning_assess_name, token, progress,
        )
        progress("  ✅ 学习评估完成，正在分析各服务器单机成本...")

        learning_assessed = azure_arm_list(
            f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
            f"/providers/Microsoft.Migrate/assessmentProjects/{project_name}"
            f"/groups/{learning_group_name}/assessments/{learning_assess_name}/assessedMachines"
            f"?api-version={AZURE_MIGRATE_API_VERSION}",
            token,
        )

        cache = learn_tier_machine_selections(learning_assessed, safe_prefix, progress)
        save_tier_cache(cache)
        progress("  ✅ 所有规模学习完成，结果已缓存到本地。")
    else:
        tier_info = cache["tiers"][str(tier)]
        progress(
            f"已命中学习缓存：规模 {_format_usd(float(tier))}，"
            f"选择 {tier_info['machine_count']} 台服务器，"
            f"预期年化 {_format_usd(tier_info['expected_annual'])}"
        )

    selected_ids = get_machine_ids_for_tier(tier, machines, safe_prefix, cache, customer_name=account_name)
    if not selected_ids:
        selected_ids = all_machine_ids
    progress(f"当前规模选定 {len(selected_ids)}/{len(all_machine_ids)} 台服务器进入最终评估")

    # V2 ARG 查询需要 OffAzure discoveryMachineArmId，不是 assessment project machine ID
    id_to_discovery = {}
    for m in machines:
        mid = m.get("id", "")
        discovery = m.get("properties", {}).get("discoveryMachineArmId", "")
        if mid and discovery:
            id_to_discovery[mid.lower()] = discovery
    selected_discovery_ids = [
        id_to_discovery.get(sid.lower(), sid) for sid in selected_ids
    ]

    # ── Step 13: 创建新版 Azure VM 评估（project-level，无需 group） ──
    progress("创建 Azure Migrate 最终评估（新版 Azure VM 类型）...")
    assessment_path = (
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Migrate/assessmentprojects/{project_name}"
        f"/assessments/{assessment_resource_name}"
        f"?api-version={AZURE_MIGRATE_V2_API_VERSION}"
    )
    assessment_body = _build_assessment_body_v2(
        import_site_id=import_site_id,
        machine_ids=selected_discovery_ids,
        target_location=target_location,
        customer_name=account_name,
    )
    settings = assessment_body.get("properties", {}).get("settings", {})
    progress(
        "  ✅ 已应用客户专属评估参数："
        f"舒适因子 {settings.get('scalingFactor')}，"
        f"折扣 {settings.get('discountPercentage')}%"
    )

    if target_location:
        progress(f"  ℹ️ 评估目标区域已设为：{target_location}（与解决方案架构一致）")
    if annual_budget is not None:
        progress(f"已解析用户预估年消耗：{_format_usd(annual_budget)}")
    else:
        progress("未解析到有效预估年消耗；评估会按 Portal 默认设置创建。")
    azure_arm_request("PUT", assessment_path, token, assessment_body)
    assessment = wait_for_assessment_complete_v2(
        subscription_id, resource_group, project_name, assessment_resource_name, token, progress
    )

    # ── Step 14: 预算调优 ──
    monthly_cost, tuning_history, budget_target_met = tune_assessment_to_budget_v2(
        subscription_id=subscription_id,
        resource_group=resource_group,
        project_name=project_name,
        assessment_name=assessment_resource_name,
        assessment_path=assessment_path,
        assessment_body=assessment_body,
        annual_budget=annual_budget,
        token=token,
        progress=progress,
        customer_name=account_name,
    )
    try:
        refresh_migrate_project_summary(subscription_id, resource_group, project_name, token)
    except Exception:
        pass

    progress("读取评估结果并下载 Azure Migrate Portal 同源 Excel 报告...")
    assessed_machines = azure_arm_list(
        f"/subscriptions/{subscription_id}/resourceGroups/{resource_group}"
        f"/providers/Microsoft.Migrate/assessmentprojects/{project_name}"
        f"/assessments/{assessment_resource_name}/assessedMachines"
        f"?api-version={AZURE_MIGRATE_V2_API_VERSION}",
        token,
    )
    excel_bytes = download_assessment_report_v2(
        subscription_id, resource_group, project_name, assessment_resource_name, token
    )
    progress(f"  ✅ Azure Migrate 导出报告已下载（{len(excel_bytes)} 字节）")
    return {
        "project_name": project_name,
        "site_name": site_name,
        "collector_name": collector_name,
        "group_name": group_name,
        "assessment_name": assessment_resource_name,
        "portal_inventory_count": portal_inventory_count,
        "migrate_project_id": migrate_project_id,
        "assessment_project_id": assessment_project_id,
        "import_site_id": import_site_id,
        "assessment": assessment,
        "assessed_machines": assessed_machines,
        "excel_bytes": excel_bytes,
        "budget_target": annual_budget,
        "budget_target_min": budget_range.get("target_min"),
        "budget_target_max": budget_range.get("target_max"),
        "budget_target_max_exclusive": budget_range.get("target_max_exclusive"),
        "budget_target_range_label": format_budget_target_range(budget_range),
        "monthly_cost": monthly_cost,
        "annualized_cost": monthly_cost * 12,
        "budget_target_met": budget_target_met,
        "tuning_history": tuning_history,
        "tier": tier,
        "selected_machine_count": len(selected_ids),
        "total_machine_count": len(all_machine_ids),
    }



def fix_assessment_excel_timestamps(
    excel_bytes: bytes,
    pov_start: datetime.date,
    pov_end: datetime.date,
) -> bytes:
    import random
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(excel_bytes))

    total_days = (pov_end - pov_start).days
    if total_days <= 0:
        total_days = 1
    random_day_offset = random.randint(1, max(total_days - 1, 1))
    created_date = pov_start + datetime.timedelta(days=random_day_offset)

    def _format_as_text(dt_val: datetime.datetime) -> str:
        hour = dt_val.hour
        ampm = "AM" if hour < 12 else "PM"
        hour_12 = hour % 12
        if hour_12 == 0:
            hour_12 = 12
        return f"{dt_val.month}/{dt_val.day}/{dt_val.year} {hour_12}:{dt_val.minute:02d}:{dt_val.second:02d} {ampm}"

    def _parse_time_from_value(orig):
        if isinstance(orig, datetime.datetime):
            return orig.hour, orig.minute, orig.second
        # Excel serial number (float/int) — fractional part encodes time-of-day
        if isinstance(orig, (int, float)):
            frac = orig % 1
            total_seconds = round(frac * 86400)
            h = total_seconds // 3600
            mi = (total_seconds % 3600) // 60
            s = total_seconds % 60
            return h, mi, s
        text = str(orig).strip()
        m = re.search(r"(\d{1,2}):(\d{2}):(\d{2})\s*(AM|PM)?", text, re.IGNORECASE)
        if m:
            h, mi, s = int(m.group(1)), int(m.group(2)), int(m.group(3))
            ampm = (m.group(4) or "").upper()
            if ampm == "PM" and h != 12:
                h += 12
            elif ampm == "AM" and h == 12:
                h = 0
            return h, mi, s
        return 2, 35, 35

    created_datetime = None

    if "Assessment_Summary" in wb.sheetnames:
        ws = wb["Assessment_Summary"]
        # Assessment_Summary 是 key-value 布局（A列=属性名, B列=值）
        for row in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and "created on" in str(cell_val).lower():
                val_col = 2
                orig = ws.cell(row=row, column=val_col).value
                if orig is not None:
                    h, mi, s = _parse_time_from_value(orig)
                    created_datetime = datetime.datetime(
                        created_date.year, created_date.month, created_date.day, h, mi, s
                    )
                    ws.cell(row=row, column=val_col).value = _format_as_text(created_datetime)
                break

    if created_datetime is None:
        now_utc = datetime.datetime.utcnow()
        created_datetime = datetime.datetime(
            created_date.year, created_date.month, created_date.day,
            now_utc.hour, now_utc.minute, now_utc.second,
        )

    if "Assessment_Properties" in wb.sheetnames:
        ws = wb["Assessment_Properties"]
        # 只修改日期部分，保留原始时分秒
        perf_end_date = created_datetime.date()
        perf_start_date = perf_end_date - datetime.timedelta(days=1)
        # 边界保护：start 不能早于 POV 开始
        if perf_start_date < pov_start:
            perf_start_date = pov_start
            perf_end_date = pov_start + datetime.timedelta(days=1)

        # 先收集 perf end 原始时间，确保 Created 时间 > Perf end 时间
        perf_end_time = None
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                prop_name = str(ws.cell(row=row, column=col).value or "").lower()
                if "performance history end" in prop_name:
                    orig = ws.cell(row=row, column=col + 1).value
                    h, mi, s = _parse_time_from_value(orig)
                    perf_end_time = (h, mi, s)
                    break
            if perf_end_time:
                break

        # 确保 created_datetime 的时间晚于 perf_end 时间（Portal 自然行为）
        if perf_end_time:
            pe_h, pe_mi, pe_s = perf_end_time
            cr_seconds = created_datetime.hour * 3600 + created_datetime.minute * 60 + created_datetime.second
            pe_seconds = pe_h * 3600 + pe_mi * 60 + pe_s
            if cr_seconds <= pe_seconds:
                # Created 应比 perf_end 晚 30-60 秒
                new_cr_seconds = pe_seconds + 34
                if new_cr_seconds >= 86400:
                    new_cr_seconds = 86400 - 1
                new_h = new_cr_seconds // 3600
                new_mi = (new_cr_seconds % 3600) // 60
                new_s = new_cr_seconds % 60
                created_datetime = datetime.datetime(
                    created_datetime.year, created_datetime.month, created_datetime.day,
                    new_h, new_mi, new_s
                )
                # 同步更新 Assessment_Summary 中的 Created on
                if "Assessment_Summary" in wb.sheetnames:
                    ws_sum = wb["Assessment_Summary"]
                    for r in range(1, ws_sum.max_row + 1):
                        cv = ws_sum.cell(row=r, column=1).value
                        if cv and "created on" in str(cv).lower():
                            ws_sum.cell(row=r, column=2).value = _format_as_text(created_datetime)
                            break

        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                prop_name = str(ws.cell(row=row, column=col).value or "").lower()
                if "performance history start" in prop_name:
                    val_col = col + 1
                    orig = ws.cell(row=row, column=val_col).value
                    h, mi, s = _parse_time_from_value(orig)
                    new_dt = datetime.datetime(perf_start_date.year, perf_start_date.month, perf_start_date.day, h, mi, s)
                    ws.cell(row=row, column=val_col).value = _format_as_text(new_dt)
                elif "performance history end" in prop_name:
                    val_col = col + 1
                    orig = ws.cell(row=row, column=val_col).value
                    h, mi, s = _parse_time_from_value(orig)
                    new_dt = datetime.datetime(perf_end_date.year, perf_end_date.month, perf_end_date.day, h, mi, s)
                    ws.cell(row=row, column=val_col).value = _format_as_text(new_dt)
                elif "created on" in prop_name:
                    val_col = col + 1
                    ws.cell(row=row, column=val_col).value = _format_as_text(created_datetime)

    out_buf = io.BytesIO()
    wb.save(out_buf)
    return out_buf.getvalue()
