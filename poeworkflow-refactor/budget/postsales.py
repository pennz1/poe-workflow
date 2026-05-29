"""售后 Azure Calculator 生成逻辑。

从售前导出的 Azure Pricing Calculator xlsx 中读取各行项目，
通过 LLM 解析 Description 中的数量信息，按缩放因子调整数量，
使年总消耗刚好超过用户设定的目标值（100%-110%）。
"""

import io
import json
import math
import random
import re
from copy import copy as _copy
from typing import Any

import openpyxl
from openpyxl.styles import Font as _Font


# ──────────────────────────────────────────────
# Step 1: 从 worksheet 中提取行项目
# ──────────────────────────────────────────────

def _find_header_row(ws):
    """找到包含 'Estimated monthly cost' 的标题行。"""
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and any(
            isinstance(v, str) and "Estimated monthly cost" in v for v in row if v
        ):
            return i
    return None


def _find_total_row(ws, hrow):
    """找到标题行之后的 Total 行。"""
    for i, row in enumerate(ws.iter_rows(min_row=hrow + 1, values_only=True), hrow + 1):
        if row and any(isinstance(v, str) and v.strip() == "Total" for v in row if v):
            return i
    return None


def extract_line_items(ws):
    """提取 worksheet 中所有数据行的 (row_idx, description, monthly_cost)。

    Returns:
        tuple: (header_row, total_row, desc_col, cost_col, items_list)
        items_list 中每项为 dict: {row, description, monthly_cost}
    """
    hrow = _find_header_row(ws)
    if hrow is None:
        return None, None, None, None, []

    trow = _find_total_row(ws, hrow)
    if trow is None:
        return None, None, None, None, []

    header_vals = [ws.cell(hrow, c).value for c in range(1, ws.max_column + 1)]

    # 找 Description 列（可能叫 "Description" 或包含 description）
    desc_col = None
    cost_col = None
    for idx, v in enumerate(header_vals):
        if isinstance(v, str):
            vl = v.strip().lower()
            if vl == "description" or "description" in vl:
                desc_col = idx + 1
            if "estimated monthly cost" in vl:
                cost_col = idx + 1

    if desc_col is None or cost_col is None:
        return None, None, None, None, []

    items = []
    for r in range(hrow + 1, trow):
        desc = ws.cell(r, desc_col).value
        cost = ws.cell(r, cost_col).value
        if desc and cost is not None:
            # cost 可能是数字或公式字符串
            try:
                cost_val = float(cost) if not isinstance(cost, (int, float)) else cost
            except (ValueError, TypeError):
                cost_val = 0.0
            items.append({
                "row": r,
                "description": str(desc).strip(),
                "monthly_cost": float(cost_val),
            })

    return hrow, trow, desc_col, cost_col, items


# ──────────────────────────────────────────────
# Step 2: LLM 解析 Description 中的数量
# ──────────────────────────────────────────────

def build_parse_user_prompt(items: list[dict]) -> str:
    """构建发送给 LLM 的 user prompt，包含所有行的 Description 和 monthly_cost。"""
    lines = []
    for item in items:
        lines.append(f"Row {item['row']}: Description=\"{item['description']}\", monthly_cost={item['monthly_cost']}")
    return "\n".join(lines)


def parse_llm_response(response_text: str) -> list[dict]:
    """解析 LLM 返回的 JSON 数组。"""
    # 去掉可能的 markdown code block
    text = response_text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)

    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        # 尝试提取 JSON 数组
        match = re.search(r"\[.*\]", text, re.DOTALL)
        if match:
            parsed = json.loads(match.group())
        else:
            return []

    if not isinstance(parsed, list):
        return []
    return parsed


# ──────────────────────────────────────────────
# Step 3: 调整数量使年总消耗达到目标
# ──────────────────────────────────────────────

def adjust_quantities(parsed_items: list[dict], items: list[dict], annual_target: float):
    """根据 LLM 解析结果和目标年消耗，计算每行的缩放因子和新数量。

    Args:
        parsed_items: LLM 解析结果列表，每项含 {row, type, quantities, adjustable}
        items: 原始数据行列表，每项含 {row, description, monthly_cost}
        annual_target: 用户期望的年消耗 USD

    Returns:
        list[dict]: 调整结果，每项含 {row, scale_factor, new_quantities, new_monthly_cost, original_desc, new_desc}
    """
    # 建立 row → item 映射
    item_map = {item["row"]: item for item in items}
    parsed_map = {p["row"]: p for p in parsed_items}

    # 当前年总消耗（原文件）
    current_annual = sum(item["monthly_cost"] * 12 for item in items)
    if current_annual <= 0:
        return []

    # 目标范围：原文件年消耗上下浮动 5%，且必须超过用户输入的 annual_target
    file_lower = current_annual * 0.95
    file_upper = current_annual * 1.05
    # 最终目标下限 = max(用户输入值, 原文件-5%)
    target_min = max(annual_target, file_lower)
    # 最终目标上限 = 原文件+5%（若用户输入超过此值，则以用户输入+5%为上限）
    if annual_target > file_upper:
        target_max = annual_target * 1.05
    else:
        target_max = file_upper

    # 可调整行的年消耗 vs 固定行
    adjustable_annual = 0.0
    fixed_annual = 0.0
    for item in items:
        p = parsed_map.get(item["row"])
        if p and p.get("adjustable", False):
            adjustable_annual += item["monthly_cost"] * 12
        else:
            fixed_annual += item["monthly_cost"] * 12

    # 可调整部分需要达到的目标
    adjustable_target = target_min - fixed_annual
    if adjustable_target <= 0:
        # 固定费用已超过目标，不需要调整
        adjustable_target = adjustable_annual  # 保持不变

    # 全局缩放因子
    if adjustable_annual > 0:
        global_k = adjustable_target / adjustable_annual
    else:
        global_k = 1.0

    # 为每行计算调整
    results = []
    running_annual = fixed_annual

    adjustable_rows = []
    for item in items:
        p = parsed_map.get(item["row"])
        if not p or not p.get("adjustable", False):
            # 不可调整行：保持原样
            results.append({
                "row": item["row"],
                "scale_factor": 1.0,
                "new_quantities": None,
                "new_monthly_cost": item["monthly_cost"],
                "original_desc": item["description"],
                "new_desc": item["description"],
                "adjustable": False,
            })
        else:
            adjustable_rows.append((item, p))

    # 为可调整行分配缩放因子（加入随机扰动使各行差异化）
    for item, p in adjustable_rows:
        # 随机扰动：在全局 k 基础上 ±15%
        k = global_k * random.uniform(0.85, 1.15)
        # 确保 k >= 0.3（不要缩得太小）
        k = max(k, 0.3)

        quantities = p.get("quantities", [])
        new_quantities = []
        for q in quantities:
            old_val = q.get("value", 1)
            if isinstance(old_val, str):
                try:
                    old_val = float(old_val.replace(",", ""))
                except ValueError:
                    old_val = 1
            new_val = max(1, round(old_val * k))
            # 确保数量有变化（售后不能和售前完全一样）
            if new_val == old_val:
                new_val = old_val + random.choice([-1, 1])
                if new_val < 1:
                    new_val = old_val + 1
            new_quantities.append({
                "old_value": old_val,
                "new_value": new_val,
                "context_text": q.get("context_text", ""),
            })

        # 计算实际缩放（基于第一个数量的比例）
        if quantities and new_quantities:
            first_old = new_quantities[0]["old_value"]
            first_new = new_quantities[0]["new_value"]
            actual_k = first_new / first_old if first_old > 0 else k
        else:
            actual_k = k

        new_cost = item["monthly_cost"] * actual_k

        # 生成新的 Description
        new_desc = _replace_quantities_in_desc(item["description"], new_quantities)

        results.append({
            "row": item["row"],
            "scale_factor": actual_k,
            "new_quantities": new_quantities,
            "new_monthly_cost": round(new_cost, 2),
            "original_desc": item["description"],
            "new_desc": new_desc,
            "adjustable": True,
        })
        running_annual += new_cost * 12

    # 微调：如果总额不在目标范围内，逐步调整可调行
    total_annual = sum(
        r["new_monthly_cost"] * 12 for r in results
    )

    # 如果低于目标，增加最大花费行的数量
    max_iterations = 50
    iteration = 0
    while total_annual < target_min and iteration < max_iterations:
        iteration += 1
        # 找到最大花费的可调整行
        adj_results = [r for r in results if r["adjustable"]]
        if not adj_results:
            break
        biggest = max(adj_results, key=lambda r: r["new_monthly_cost"])
        # 增加 5%
        biggest["new_monthly_cost"] = round(biggest["new_monthly_cost"] * 1.05, 2)
        if biggest["new_quantities"]:
            for q in biggest["new_quantities"]:
                q["new_value"] = max(1, round(q["new_value"] * 1.05))
            biggest["new_desc"] = _replace_quantities_in_desc(
                biggest["original_desc"], biggest["new_quantities"]
            )
            # 重算 scale_factor
            if biggest["new_quantities"][0]["old_value"] > 0:
                biggest["scale_factor"] = (
                    biggest["new_quantities"][0]["new_value"]
                    / biggest["new_quantities"][0]["old_value"]
                )
        total_annual = sum(r["new_monthly_cost"] * 12 for r in results)

    # 如果超出太多（>110%），缩小最大行
    iteration = 0
    while total_annual > target_max and iteration < max_iterations:
        iteration += 1
        adj_results = [r for r in results if r["adjustable"]]
        if not adj_results:
            break
        biggest = max(adj_results, key=lambda r: r["new_monthly_cost"])
        biggest["new_monthly_cost"] = round(biggest["new_monthly_cost"] * 0.95, 2)
        if biggest["new_quantities"]:
            for q in biggest["new_quantities"]:
                q["new_value"] = max(1, round(q["new_value"] * 0.95))
            biggest["new_desc"] = _replace_quantities_in_desc(
                biggest["original_desc"], biggest["new_quantities"]
            )
            if biggest["new_quantities"][0]["old_value"] > 0:
                biggest["scale_factor"] = (
                    biggest["new_quantities"][0]["new_value"]
                    / biggest["new_quantities"][0]["old_value"]
                )
        total_annual = sum(r["new_monthly_cost"] * 12 for r in results)

    return results


def _replace_quantities_in_desc(desc: str, new_quantities: list[dict]) -> str:
    """在 Description 中替换数量值。

    使用 context_text 精确定位要替换的数字。
    """
    new_desc = desc
    for q in new_quantities:
        context = q.get("context_text", "")
        old_val = q.get("old_value", 0)
        new_val = q.get("new_value", 0)

        if not context or old_val == new_val:
            continue

        # 将数字格式化为与原文一致的形式
        old_str = _format_number_like_original(old_val, desc)
        new_str = _format_number(new_val, old_val, desc)

        if context in new_desc:
            # 在 context 范围内替换数字
            old_context = context
            new_context = context.replace(old_str, new_str, 1)
            if old_context != new_context:
                new_desc = new_desc.replace(old_context, new_context, 1)
        else:
            # fallback：直接替换第一个出现的数字
            new_desc = new_desc.replace(old_str, new_str, 1)

    return new_desc


def _format_number_like_original(value, original_text: str) -> str:
    """检测原文中的数字格式（带逗号 vs 不带）并返回对应格式。"""
    # 检查原文中该数值是否带逗号
    val_with_comma = f"{int(value):,}"
    val_without_comma = str(int(value))

    if val_with_comma in original_text:
        return val_with_comma
    return val_without_comma


def _format_number(new_value, old_value, original_text: str) -> str:
    """根据原文中旧值的格式来格式化新值。"""
    old_with_comma = f"{int(old_value):,}"
    if old_with_comma in original_text:
        return f"{int(new_value):,}"
    return str(int(new_value))


# ──────────────────────────────────────────────
# Step 4: 将调整写回 worksheet
# ──────────────────────────────────────────────

def apply_changes(ws, hrow, trow, desc_col, cost_col, adjustments: list[dict]):
    """将调整后的 Description 和 monthly_cost 写回 worksheet。"""
    for adj in adjustments:
        r = adj["row"]
        # 写入新的 Description
        ws.cell(r, desc_col).value = adj["new_desc"]
        # 写入新的 monthly cost
        ws.cell(r, cost_col).value = round(adj["new_monthly_cost"], 2)

    # 重新计算 Total 行
    total = sum(adj["new_monthly_cost"] for adj in adjustments)
    ws.cell(trow, cost_col).value = round(total, 2)


# ──────────────────────────────────────────────
# Step 5: 主流程函数
# ──────────────────────────────────────────────

def process_postsales_excel(
    file_bytes: bytes,
    annual_target: float,
    llm_call_fn,
    system_prompt: str,
    progress_callback=None,
) -> tuple[bytes, dict[str, Any]]:
    """处理售后 Excel 的主入口。

    Args:
        file_bytes: 上传的 xlsx 文件字节
        annual_target: 目标年消耗 USD
        llm_call_fn: LLM 调用函数 (system_prompt, user_prompt) -> str
        system_prompt: LLM system prompt
        progress_callback: 可选的进度回调 (message: str) -> None

    Returns:
        tuple: (output_xlsx_bytes, summary_info)
    """
    def _progress(msg):
        if progress_callback:
            progress_callback(msg)

    _progress("正在读取 Excel...")
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))

    all_results = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        hrow, trow, desc_col, cost_col, items = extract_line_items(ws)

        if not items:
            all_results[sheet_name] = {"status": "skipped", "reason": "未找到有效数据行"}
            continue

        _progress(f"正在解析 [{sheet_name}] 中 {len(items)} 个行项目...")

        # 调用 LLM 解析
        user_prompt = build_parse_user_prompt(items)
        try:
            llm_response = llm_call_fn(system_prompt, user_prompt)
            parsed_items = parse_llm_response(llm_response)
        except Exception as e:
            all_results[sheet_name] = {"status": "error", "reason": f"LLM 解析失败: {e}"}
            continue

        if not parsed_items:
            all_results[sheet_name] = {"status": "error", "reason": "LLM 返回结果为空或格式错误"}
            continue

        _progress(f"正在调整 [{sheet_name}] 的数量以达到年消耗目标...")

        # 调整数量
        adjustments = adjust_quantities(parsed_items, items, annual_target)

        # 写回
        apply_changes(ws, hrow, trow, desc_col, cost_col, adjustments)

        # 统计
        new_annual = sum(a["new_monthly_cost"] * 12 for a in adjustments)
        old_annual = sum(item["monthly_cost"] * 12 for item in items)
        adjusted_count = sum(1 for a in adjustments if a["adjustable"])

        all_results[sheet_name] = {
            "status": "success",
            "old_annual": round(old_annual, 2),
            "new_annual": round(new_annual, 2),
            "target": annual_target,
            "adjusted_rows": adjusted_count,
            "total_rows": len(items),
        }

    _progress("正在生成输出文件...")
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)

    return out_buf.getvalue(), all_results
