"""Tests for new features: SVG generation, timestamp fix, prompt updates."""
import io
import datetime
import unittest.mock
import sys
import os
import importlib

# Ensure real modules are used (test_tier_algorithm.py may mock them at module level)
_mocked_prefixes = ("docx", "openai", "streamlit", "msal")
for mod_name in list(sys.modules.keys()):
    if any(mod_name == p or mod_name.startswith(p + ".") for p in _mocked_prefixes):
        if hasattr(sys.modules[mod_name], '_mock_name') or hasattr(sys.modules[mod_name], '_mock_children'):
            del sys.modules[mod_name]

sys.path.insert(0, os.path.dirname(__file__))

# Force re-import of app if it was imported with mocked dependencies
if "app" in sys.modules:
    importlib.reload(sys.modules["app"])


def test_extract_svg_from_code_block():
    from app import _extract_svg_from_response
    text = '```svg\n<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 1200 800">\n<rect/>\n</svg>\n```'
    result = _extract_svg_from_response(text)
    assert "<svg" in result and "</svg>" in result


def test_extract_svg_raw():
    from app import _extract_svg_from_response
    text = '<svg xmlns="http://www.w3.org/2000/svg"><rect/></svg>'
    result = _extract_svg_from_response(text)
    assert "<svg" in result


def test_generate_svg_architecture_mock():
    from app import generate_svg_architecture
    mock_svg = '<svg xmlns="http://www.w3.org/2000/svg"><rect/></svg>'
    with unittest.mock.patch("app.call_azure_openai", return_value=mock_svg):
        svg = generate_svg_architecture(
            "## 二、架构概览\nSome content\n## 三、其他\nOther",
            "TestCo",
        )
    assert svg is not None
    assert "<svg" in svg


def test_generate_svg_extracts_relevant_sections():
    from app import generate_svg_architecture
    content = (
        "# Title\n"
        "## 一、概述\nIntro\n"
        "## 二、架构概览\nArchitecture stuff\n"
        "## 三、技术实现\nTech details\n"
        "## 五、资源列表\nResources\n"
        "## 六、部署\nDeploy\n"
        "## 七、安全\nSecurity\n"
        "## 八、监控\nMonitor\n"
    )
    mock_svg = '<svg xmlns="http://www.w3.org/2000/svg"><rect/></svg>'
    with unittest.mock.patch("app.call_azure_openai", return_value=mock_svg) as mock_call:
        generate_svg_architecture(content, "TestCo")
        user_prompt = mock_call.call_args[0][1]
        # Should include chapters 2, 5, 6, 7, 8 but NOT 1, 3
        assert "架构概览" in user_prompt
        assert "资源列表" in user_prompt
        assert "概述" not in user_prompt
        assert "技术实现" not in user_prompt


def test_fix_assessment_excel_timestamps():
    from openpyxl import Workbook, load_workbook
    from app import fix_assessment_excel_timestamps

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Assessment_Summary"
    ws1["A1"] = "Name"
    ws1["B1"] = "Created on (UTC)"
    ws1["A2"] = "Test"
    ws1["B2"] = datetime.datetime(2020, 1, 1)

    ws2 = wb.create_sheet("Assessment_Properties")
    ws2["A1"] = "Property Name"
    ws2["B1"] = "Value"
    ws2["A2"] = "Performance history start time"
    ws2["B2"] = datetime.datetime(2020, 1, 1)
    ws2["A3"] = "Performance history end time"
    ws2["B3"] = datetime.datetime(2020, 1, 31)

    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    pov_start = datetime.date(2025, 3, 1)
    pov_end = datetime.date(2025, 3, 15)
    fixed = fix_assessment_excel_timestamps(excel_bytes, pov_start, pov_end)

    wb2 = load_workbook(io.BytesIO(fixed))
    created = wb2["Assessment_Summary"]["B2"].value
    assert pov_start <= created.date() <= pov_end

    perf_start = wb2["Assessment_Properties"]["B2"].value
    perf_end = wb2["Assessment_Properties"]["B3"].value
    assert pov_start <= perf_start.date() <= pov_end
    assert pov_start <= perf_end.date() <= pov_end
    assert perf_start <= perf_end


def test_solution_prompt_uses_latest_models():
    from app import SOLUTION_SYSTEM_PROMPT
    assert "GPT-5.5" in SOLUTION_SYSTEM_PROMPT
    assert "GPT-5.4" in SOLUTION_SYSTEM_PROMPT
    assert "Azure AI Search" in SOLUTION_SYSTEM_PROMPT
    assert "Azure AI Document Intelligence" in SOLUTION_SYSTEM_PROMPT
    assert "Azure AI Content Safety" in SOLUTION_SYSTEM_PROMPT
    # Should NOT have standalone GPT-4o as recommended model
    # (it may appear in "禁止使用已过时的 GPT-4o" context, which is fine)


def test_pov_prompt_uses_latest_models():
    from app import POV_SYSTEM_PROMPT
    assert "GPT-5.4-mini" in POV_SYSTEM_PROMPT
    assert "GPT-5.4" in POV_SYSTEM_PROMPT


def test_create_solution_docx_with_svg():
    from app import create_solution_docx
    # Create a minimal PNG (1x1 white pixel)
    import struct
    import zlib

    def _make_minimal_png():
        sig = b'\x89PNG\r\n\x1a\n'
        # IHDR
        ihdr_data = struct.pack('>IIBBBBB', 1, 1, 8, 2, 0, 0, 0)
        ihdr_crc = zlib.crc32(b'IHDR' + ihdr_data) & 0xffffffff
        ihdr = struct.pack('>I', 13) + b'IHDR' + ihdr_data + struct.pack('>I', ihdr_crc)
        # IDAT
        raw = zlib.compress(b'\x00\xff\xff\xff')
        idat_crc = zlib.crc32(b'IDAT' + raw) & 0xffffffff
        idat = struct.pack('>I', len(raw)) + b'IDAT' + raw + struct.pack('>I', idat_crc)
        # IEND
        iend_crc = zlib.crc32(b'IEND') & 0xffffffff
        iend = struct.pack('>I', 0) + b'IEND' + struct.pack('>I', iend_crc)
        return sig + ihdr + idat + iend

    content = (
        "# Test Solution\n"
        "## 一、概述\nIntro\n"
        "## 二、解决方案架构概览\nArchitecture\n"
        "## 三、技术实现\nDetails\n"
    )
    png_bytes = _make_minimal_png()
    docx_bytes = create_solution_docx(content, "TestCo", svg_png_bytes=png_bytes)
    assert len(docx_bytes) > 0
    # Verify it's a valid docx (ZIP)
    import zipfile
    assert zipfile.is_zipfile(io.BytesIO(docx_bytes))


if __name__ == "__main__":
    import pytest
    pytest.main([__file__, "-v"])
